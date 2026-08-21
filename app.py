"""
Prospect Drafter — web app.

Three workflows, chosen at the top of the page and configured under "workflows" in
config.json. Each workflow declares a `mode`, and the app branches on that rather than on
which workflow is selected:

  mode "sequence" — an AI-written email per prospect, with the sequence driven off the
  sheet's Status column, which is also what gets written back. Two of these:

    In-bound Leads  people who came to WLCC. Sheet "Inbound Leads", Status in column C.
    Cold Call       cold outreach. Sheet "Cold Database in Work", Status in column A,
                    and rows are picked by WLCC Fit Score first.

  Four steps: upload the list, pick which email in the sequence, read the drafts, put them
  in Outlook. Stages come from config, not from a fixed count:
    1  first email
    2  follow-up, ~7 days later, only if they didn't reply
    3  final follow-up, ~7 days after that

  mode "batch" — Batch Email. One email the user has already written, merged with a list.
  No research and no AI at all; see batch_email.py.

Nothing here is workflow-specific in code: the sheet, the columns, the templates and
whether there is a score to filter on all come from config.

Nothing is ever sent, in any workflow. Drafts go into Outlook unsent.

Deployed on Streamlit Community Cloud as a private app.
"""

from __future__ import annotations

import base64
import importlib.util
import io
import json
import os
import re
import sys
import time
import urllib.parse
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import streamlit as st

import batch_email as be_lib
import prospect_drafter as pd_lib
import xlsx_patch

APP_DIR = Path(__file__).parent
CONFIG_PATH = APP_DIR / "config.json"

# Kept separate from pd_lib.GRAPH_SCOPES (mail-only) on purpose: OneDrive access is an
# independent, optional sign-in. If the Entra app registration doesn't have Files.ReadWrite
# added yet, only the OneDrive connect button fails — the working Outlook connection is
# never put at risk by bundling an unapproved scope into it.
ONEDRIVE_SCOPES = "offline_access Files.ReadWrite User.Read"

st.set_page_config(page_title="Prospect Drafter", page_icon="✉️", layout="centered")

# Streamlit re-runs this file on every click but keeps imported modules cached in
# sys.modules, and its hot-reload of a local module alongside the app is not reliable. After
# an update that touches both files, a running server can end up executing the new app.py
# against the old batch_email - which surfaces as AttributeError four steps into a batch,
# with a traceback nobody outside this file can act on. Checked once, here, so it says what
# to do instead.
_NEEDED = ("find_links", "linkify", "delinkify", "plain_links", "build_drafts", "read_table")
_missing = [name for name in _NEEDED if not hasattr(be_lib, name)]
if _missing:
    st.error(
        "**The app has been updated and needs restarting.**\n\n"
        "It is running new code against a part of itself that is still the old version, so "
        "some things would fail halfway through. Nothing is lost — close the black window "
        "this app was started from and run **run-local.cmd** again.\n\n"
        f"(For Mo: batch_email is stale in this process — missing {', '.join(_missing)}.)"
    )
    st.stop()


# ---------------------------------------------------------------------------
# Config and secrets
# ---------------------------------------------------------------------------

@st.cache_data
def load_config() -> dict:
    return pd_lib.deep_merge(pd_lib.DEFAULT_CONFIG, json.loads(CONFIG_PATH.read_text("utf-8")))


def secret(name: str, default: str = "") -> str:
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return os.environ.get(name, default)


def get_config(workflow: str | None = None) -> dict:
    cfg = json.loads(json.dumps(load_config()))
    if workflow:
        cfg = pd_lib.apply_workflow(cfg, workflow)
    if secret("GRAPH_CLIENT_ID"):
        cfg["outlook"]["graph_client_id"] = secret("GRAPH_CLIENT_ID")
    if secret("GRAPH_TENANT_ID"):
        cfg["outlook"]["graph_tenant_id"] = secret("GRAPH_TENANT_ID")
    for key in ("your_name", "your_company", "signature_html"):
        val = st.session_state.get(f"sender_{key}")
        if val:
            cfg["sender"][key] = val
    return cfg


@st.cache_data(ttl=300, show_spinner=False)
def cached_ai_key_check(key: str, model: str, base_url: str | None) -> dict:
    """Cached so the live check runs once every 5 minutes, not on every widget click.

    The key value is part of the cache's own lookup key, so fixing a broken key in Secrets
    and reloading the app is checked fresh immediately rather than reusing a stale failure.
    """
    return pd_lib.check_ai_key(key, model, base_url)


# ---------------------------------------------------------------------------
# HTML <-> plain text, so nobody has to edit HTML tags
# ---------------------------------------------------------------------------
# Drafts are held as plain text everywhere in the app, because that is what a person can read
# and edit in a box. Links are the one thing plain text cannot carry on its own, so they are
# written down as [text](url) on the way in and turned back into real anchors on the way out.
# See the Links section of batch_email.py.

def html_to_plain(html: str) -> str:
    text = be_lib.delinkify(html or "")
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p\s*>", "\n\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    for a, b in (("&nbsp;", " "), ("&amp;", "&"), ("&lt;", "<"), ("&gt;", ">"),
                 ("&#39;", "'"), ("&quot;", '"')):
        text = text.replace(a, b)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def plain_to_html(text: str) -> str:
    blocks = [b.strip() for b in (text or "").split("\n\n") if b.strip()]
    return "\n".join(
        "<p>" + be_lib.linkify(pd_lib.escape_html(b)) + "</p>" for b in blocks
    )


# ---------------------------------------------------------------------------
# Spreadsheet
# ---------------------------------------------------------------------------

def save_upload(uploaded) -> Path:
    tmp = APP_DIR / "_uploads"
    tmp.mkdir(exist_ok=True)
    path = tmp / uploaded.name
    path.write_bytes(uploaded.getbuffer())
    return path


def _open_sheet(source: Path, cfg: dict):
    from openpyxl import load_workbook, Workbook

    if source.suffix.lower() == ".csv":
        import csv
        wb = Workbook()
        ws = wb.active
        for r in csv.reader(source.open(encoding="utf-8-sig")):
            ws.append(r)
        return wb, ws

    wb = load_workbook(source)
    want = cfg["spreadsheet"].get("sheet_name")
    ws = wb[want] if want in wb.sheetnames else wb[wb.sheetnames[0]]
    return wb, ws


def spreadsheet_with_progress(source: Path, updates: dict[str, dict], cfg: dict) -> bytes:
    """Copy of the sheet with Touches and the contact dates brought up to date.

    `updates` maps lowercase email -> {"touches": int, "date": date}.
    Columns that don't exist yet are appended to the header row.
    """
    wb, ws = _open_sheet(source, cfg)
    colmap = cfg["spreadsheet"]["columns"]
    hdr_row = int(cfg["spreadsheet"].get("header_row", 1))

    headers: dict[str, int] = {}
    for cell in ws[hdr_row]:
        if cell.value is not None and str(cell.value).strip():
            headers[str(cell.value).strip().lower()] = cell.column

    def column_for(field: str, fallback_title: str) -> int | None:
        title = str(colmap.get(field) or fallback_title).strip()
        if not title:
            return None
        idx = headers.get(title.lower())
        if idx is None:
            idx = ws.max_column + 1
            ws.cell(row=hdr_row, column=idx).value = title
            headers[title.lower()] = idx
        return idx

    email_col = headers.get(str(colmap.get("email") or "Email").strip().lower())
    if not email_col:
        raise RuntimeError("Could not find the Email column to update.")

    touch_col = column_for("touches", "Touches")
    first_col = column_for("first_contact_date", "First Contact Date")
    last_col = column_for("last_contact_date", "Last Contact Date")

    for row in range(hdr_row + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=email_col).value
        if not v:
            continue
        u = updates.get(str(v).strip().lower())
        if not u:
            continue
        when = u["date"].strftime("%Y-%m-%d")
        if touch_col:
            ws.cell(row=row, column=touch_col).value = u["touches"]
        if first_col and not ws.cell(row=row, column=first_col).value:
            ws.cell(row=row, column=first_col).value = when
        if last_col:
            ws.cell(row=row, column=last_col).value = when

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def workbook_sheets(source: Path) -> tuple[list[str], str | None]:
    """Visible sheet names, plus whichever sheet Excel had active when it was saved.

    Hidden sheets are left out: this workbook keeps its dropdown lists on a hidden
    "Sheet Info" tab, and offering that as somewhere to read prospects from is noise.
    """
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True)
    try:
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        try:
            active = wb.active.title if wb.active is not None else None
        except Exception:  # noqa: BLE001 - a stale active index shouldn't stop the upload
            active = None
        return (visible or list(wb.sheetnames)), active
    finally:
        wb.close()


def sheet_selector(source: Path, cfg: dict) -> str | None:
    """Show which sheet the app is about to read, and let it be changed.

    Each workflow names the sheet it expects, so the normal case is that this is already
    correct and the user just sees it confirmed. Returns None for CSV, which has no sheets.
    """
    if source.suffix.lower() == ".csv":
        st.caption("A CSV has just the one sheet.")
        return None

    try:
        names, excel_active = workbook_sheets(source)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not open that workbook: {exc}")
        st.stop()

    want = str(cfg["spreadsheet"].get("sheet_name") or "").strip()
    label = cfg.get("workflow", {}).get("label", "this workflow")
    configured = pd_lib.find_sheet(names, want)
    default = pd_lib.resolve_sheet_name(names, want, excel_active)

    st.markdown("**Active sheet**")
    chosen = st.selectbox(
        "Active sheet",
        options=names,
        index=names.index(default),
        label_visibility="collapsed",
        format_func=lambda n: f"{n}  ·  active in Excel" if n == excel_active else n,
        help=(
            f"The {label} workflow reads “{want}”. Change this only if you keep that data "
            "on a differently named sheet."
            if want else "Which sheet holds the prospects."
        ),
    )

    if want and not configured:
        st.warning(
            f"This workbook has no **{want}** sheet, which is the one the **{label}** "
            f"workflow expects. Reading **{chosen}** instead — check that's right, or "
            "switch workflow."
        )
    elif configured and chosen != configured:
        st.warning(
            f"Reading **{chosen}**, not the **{configured}** sheet that **{label}** "
            "normally uses. The column names have to match, or nothing will be found."
        )
    else:
        st.caption(f"Reading **{chosen}** — the sheet the {label} workflow expects.")

    return chosen


# ---------------------------------------------------------------------------
# Cold Call: fit score filter and Status write-back
# ---------------------------------------------------------------------------

def fit_score_filter(prospects: list, cfg: dict) -> list:
    """Let the user pick which WLCC Fit Scores to work, and keep only those rows.

    This is step one of the Cold Call workflow: the score in column B is a human
    judgement about how well a company fits WLCC, and the whole point is to work the
    good ones first rather than the whole 800-row database.
    """
    spec = cfg["workflow"]["fit_score"]
    label = spec.get("label", "Fit score")
    present = pd_lib.fit_score_values(prospects, cfg)
    counts = Counter(p.fit_score for p in prospects)
    unscored = counts.get("", 0)

    if not present:
        st.error(
            f'No "{label}" values found in this sheet. Check the sheet and the '
            f"`fit_score` column name in the Cold Call config."
        )
        st.stop()

    st.markdown(f"**Which {label}s are you working?**")
    default = [s for s in spec.get("default", []) if s in present] or present[:1]
    chosen = st.multiselect(
        label, options=present, default=default,
        format_func=lambda s: f"{s} — {counts[s]} companies",
        label_visibility="collapsed",
        help="Pick one or more. 5 is the best fit. Only these rows go any further.",
    )
    if not chosen:
        st.warning(f"Pick at least one {label} to carry on.")
        st.stop()

    kept = [p for p in prospects if p.fit_score in set(chosen)]
    st.caption(
        f"**{len(kept)}** companies scored {', '.join(chosen)}."
        + (f" {unscored} rows have no score yet and are left out." if unscored else "")
    )
    if not kept:
        st.warning("No rows have those scores.")
        st.stop()
    return kept


def status_writeback_edits(drafts: list[dict], cfg: dict, today: date) -> dict[int, dict[str, object]]:
    """The cell edits that record this batch: Status, and the two contact dates.

    Keyed by worksheet row rather than by email address, because the cold database has
    duplicate and blank email cells and the row number is exact.
    """
    wb = cfg["workflow"]["writeback"]
    status_col = str(wb.get("status_column") or "").strip()
    first_col = str(wb.get("first_contact_column") or "").strip()
    last_col = str(wb.get("last_contact_column") or "").strip()

    edits: dict[int, dict[str, object]] = {}
    for d in drafts:
        cells: dict[str, object] = {}
        if status_col and d.get("next_status"):
            cells[status_col] = d["next_status"]
        # First Contact Date is written once and then left alone.
        if first_col and d.get("needs_first_date"):
            cells[first_col] = today
        if last_col:
            cells[last_col] = today
        if cells:
            edits[int(d["row"])] = cells
    return edits


def status_writeback(source: Path, drafts: list[dict], cfg: dict, today: date) -> bytes:
    """The workbook with this batch's Status and dates written in.

    Uses xlsx_patch rather than openpyxl on purpose. See that module's docstring: an
    openpyxl save of the WLCC master workbook drops the dropdowns and most of the
    conditional formatting on the other sheets.
    """
    if source.suffix.lower() not in (".xlsx", ".xlsm"):
        raise RuntimeError(
            "Writing the Status column back needs the Excel file itself (.xlsx), not a CSV."
        )
    edits = status_writeback_edits(drafts, cfg, today)
    if not edits:
        raise RuntimeError("Nothing to write back.")
    return xlsx_patch.patch_workbook_bytes(
        source.read_bytes(),
        cfg["spreadsheet"]["sheet_name"],
        edits,
        int(cfg["spreadsheet"].get("header_row", 1)),
    )


# ---------------------------------------------------------------------------
# Outlook
# ---------------------------------------------------------------------------

def outlook_web_link(to: str, subject: str, body_plain: str) -> str:
    q = urllib.parse.urlencode(
        {"to": to, "subject": subject, "body": body_plain}, quote_via=urllib.parse.quote
    )
    return f"https://outlook.office.com/mail/deeplink/compose?{q}"


# How long a compose deeplink may get before it is not worth trying. The whole email travels
# in the URL's query string, and Microsoft's endpoints sit behind the usual 2048-byte
# query-string ceiling — over that the link doesn't degrade gracefully, it opens an error page
# or a silently truncated message. So this stays where it is: about 1150 characters of email.
# Anything longer gets a draft file instead, which has no length limit at all.
OUTLOOK_LINK_MAX = 1900


def build_eml(d: dict, cfg: dict) -> bytes:
    """One draft as a .eml file. Outlook opens it as an unsent, editable message.

    X-Unsent is what makes the difference between Outlook showing this as a received message
    and showing it as a draft with a Send button.
    """
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["Subject"] = d["subject"]
    msg["To"] = d["to"]
    msg["X-Unsent"] = "1"
    html = plain_to_html(d["body"]) + "\n" + cfg["sender"].get("signature_html", "")
    # The plain-text half of the message cannot hold an anchor, so links are flattened to
    # "text (url)" there rather than shown as [text](url).
    msg.set_content(be_lib.plain_links(d["body"]))
    msg.add_alternative(f"<html><body>{html}</body></html>", subtype="html")
    return msg.as_bytes()


def eml_filename(d: dict, index: int | None = None) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", d.get("to") or f"row{d.get('row', 0)}")[:60]
    return f"{index:03d}_{safe}.eml" if index is not None else f"{safe}.eml"


def build_eml_zip(drafts: list[dict], cfg: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for i, d in enumerate(drafts, 1):
            z.writestr(eml_filename(d, i), build_eml(d, cfg))
    return buf.getvalue()


@st.cache_data(ttl=120, show_spinner=False)
def outlook_on_this_computer() -> bool:
    """True if the app can hand drafts to classic Outlook running on this same machine.

    This is the route that does what people expect: the draft appears in the Outlook Drafts
    folder, unsent, with no sign-in, no download and no limit on how long the email is. It
    only exists when the app is being run locally — on Streamlit Cloud there is no Outlook on
    the server, so Graph remains the answer there.

    Deliberately does not call Dispatch to find out. That would launch Outlook just to answer
    a question nobody asked, on every rerun. Checking that pywin32 is importable and that
    Outlook has registered its ProgID establishes the same thing without side effects; if the
    call then fails anyway, the button reports why.
    """
    if sys.platform != "win32" or importlib.util.find_spec("win32com") is None:
        return False
    import winreg

    for root in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
        try:
            with winreg.OpenKey(root, r"SOFTWARE\Classes\Outlook.Application\CLSID"):
                return True
        except OSError:
            continue
    return False


def draft_label(d: dict) -> str:
    name = d.get("name") or d.get("to") or "(no name)"
    company = (d.get("company") or "").strip()
    return f"{name} — {company}" if company else name


def deliver_drafts(approved: list[dict], cfg: dict, file_stem: str, on_pushed=None) -> None:
    """The last step of every workflow: get the approved drafts into Outlook, unsent.

    Shared, because how the drafts were written has nothing to do with how they get into
    Outlook, and a batch of a hundred needs exactly the same three routes as a batch of
    three. `on_pushed(done, failed)` lets a workflow record the outcome — Batch Email keeps
    a log, the sequence workflows write the spreadsheet back instead.

    Every draft passes one final check here: any {{placeholder}} still standing means it was
    never filled in, and that draft is held back. Batch Email blocks those rows long before
    this point, but a subject line edited by hand in the review step can reintroduce one,
    and this is the single place all three routes go through.
    """
    holes = [d for d in approved if be_lib.has_placeholder(d.get("subject", ""))
             or be_lib.has_placeholder(d.get("body", ""))]
    if holes:
        st.error(
            f"**{len(holes)} draft(s) still have a `{{{{...}}}}` gap in them** and are held "
            "back until it is filled in or the draft is unticked: "
            + ", ".join(draft_label(d) for d in holes[:8])
            + ("..." if len(holes) > 8 else "")
        )
        held = {id(d) for d in holes}
        approved = [d for d in approved if id(d) not in held]
        if not approved:
            return

    tabs = st.tabs(["Straight into Outlook", "One at a time", "Download files"])

    with tabs[0]:
        # Two ways to reach Drafts, and which one is available depends on where the app is
        # running rather than on anything the user did. Outlook on this same computer needs no
        # sign-in and no setup at all, so it wins when it is there; on Streamlit Cloud there is
        # no Outlook on the server and Graph is the only way.
        def push_all(create, label: str) -> None:
            done, failed = [], []
            bar = st.progress(0.0)
            for i, d in enumerate(approved, 1):
                entry = {"to": d["to"], "subject": d["subject"],
                         "body_html": plain_to_html(d["body"])}
                try:
                    create(entry)
                    done.append(d)
                except Exception as exc:  # noqa: BLE001
                    failed.append((d, str(exc)))
                bar.progress(i / len(approved))
            if done:
                st.success(
                    f"{len(done)} draft(s) created. {label} Nothing has been sent — each one "
                    "is waiting for you to read it and press Send."
                )
            for d, err in failed:
                st.error(f"{d['to']}: {err}")
            if on_pushed:
                on_pushed(done, failed)

        if outlook_on_this_computer():
            st.caption(
                "Outlook is on this computer, so drafts go straight into your **Drafts** "
                "folder — no sign-in, no downloading, and no limit on how long the email is. "
                "Outlook will start up if it isn't already running."
            )
            if st.button(f"Create {len(approved)} drafts in Outlook", type="primary",
                         use_container_width=True, key="push_local_outlook"):
                push_all(lambda e: pd_lib.push_com(e, cfg), "Look in **Drafts** in Outlook.")
            if cfg["outlook"].get("graph_client_id") and st.session_state.get("graph_token"):
                st.caption("You're also signed in to Outlook online:")
                if st.button("Create them in Outlook online instead",
                             use_container_width=True, key="push_graph_alt"):
                    push_all(
                        lambda e: pd_lib.push_graph(e, cfg, st.session_state["graph_token"]),
                        "Look in **Drafts**.",
                    )
        elif not cfg["outlook"].get("graph_client_id"):
            st.info(
                "Not switched on yet. On this page it needs a one-off setup by IT — but if "
                "you run the app on your own computer with Outlook installed, it works with "
                "no setup at all. Until then use **One at a time**."
            )
        elif not st.session_state.get("graph_token"):
            st.info("Connect to Outlook in the sidebar first.")
        elif st.button(f"Create {len(approved)} drafts in Outlook", type="primary",
                       use_container_width=True, key="push_graph_only"):
            push_all(
                lambda e: pd_lib.push_graph(e, cfg, st.session_state["graph_token"]),
                "Look in **Drafts**.",
            )

    with tabs[1]:
        st.caption(
            "Click a name to open a new Outlook message with everything filled in. Check it, "
            "then press Send. Works with Outlook in the browser."
        )
        # A longer email cannot travel in a URL - see OUTLOOK_LINK_MAX. Rather than dump the
        # wording on screen to be copied by hand, that draft is offered as a file: saving it
        # and opening it puts the same message in front of you, formatting and links intact,
        # with no length limit. Which of the two you get depends only on how long the email
        # is, so a batch of a longer template lands here for every row while the AI
        # workflows, capped at about 130 words, almost never do.
        long_ones = 0
        for i, d in enumerate(approved, 1):
            # A compose deeplink carries plain text only, so the same flattening applies.
            body = be_lib.plain_links(
                d["body"] + "\n\n" + html_to_plain(cfg["sender"].get("signature_html", ""))
            )
            link = outlook_web_link(d["to"], d["subject"], body)
            if len(link) <= OUTLOOK_LINK_MAX:
                st.link_button(f"✉️ {draft_label(d)}", link, use_container_width=True)
                continue

            long_ones += 1
            st.download_button(
                f"✉️ {draft_label(d)} — save as a draft, then open it",
                data=build_eml(d, cfg),
                file_name=eml_filename(d),
                mime="message/rfc822",
                use_container_width=True,
                key=f"one_eml_{i}_{d.get('row', i)}",
            )
            with st.expander(f"…or copy {d.get('name') or d['to']} by hand"):
                st.code(f"To: {d['to']}\nSubject: {d['subject']}\n\n{body}", language=None)

        if long_ones:
            lead = ("**One of these is too long to send by link.** " if long_ones == 1 else
                    f"**{long_ones} of these are too long to send by link.** ")
            why = ("The whole email has to fit inside the web address for that to work, and "
                   "yours is longer than that allows — nothing is wrong with it.\n\n")
            # Where Outlook is on this machine there is no reason to make anyone save and
            # open a file, so say so plainly rather than explaining the file.
            if outlook_on_this_computer():
                st.info(
                    lead + why
                    + "**Use “Straight into Outlook” instead** — it puts every one of them "
                    "into your Drafts folder directly, whatever the length. The button here "
                    "saves the draft as a file you would then have to open, which is only "
                    "worth doing if you want the file itself."
                )
            else:
                st.info(
                    lead + why
                    + "Use the button instead: it saves the draft as a file, and opening that "
                    "file puts the message in Outlook complete with its formatting and links, "
                    "ready to read and send. **Straight into Outlook** does the same for all "
                    "of them at once, with no length limit, once it has been switched on."
                )

    with tabs[2]:
        st.caption("Email files you can keep or forward. Not importable into Outlook on the web.")
        st.download_button(
            "Download the drafts as .eml files", data=build_eml_zip(approved, cfg),
            file_name=f"{file_stem}.zip", mime="application/zip",
            use_container_width=True,
        )


def graph_has_reply(token: str, email: str, since: date) -> bool:
    """True if any message from this address arrived on or after `since`."""
    import requests

    address = email.replace("'", "''")          # OData escaping
    params = {
        "$filter": (
            f"from/emailAddress/address eq '{address}' "
            f"and receivedDateTime ge {since.isoformat()}T00:00:00Z"
        ),
        "$select": "id",
        "$top": "1",
    }
    r = requests.get(
        f"{pd_lib.GRAPH_BASE}/me/messages",
        headers={"Authorization": f"Bearer {token}"},
        params=params,
        timeout=30,
    )
    if r.status_code != 200:
        raise RuntimeError(f"Mailbox check failed ({r.status_code}): {r.text[:200]}")
    return bool(r.json().get("value"))


def device_code_start(cfg: dict, scope: str | None = None) -> dict:
    import requests
    r = requests.post(
        f"{pd_lib.LOGIN_BASE}/{cfg['outlook'].get('graph_tenant_id') or 'organizations'}"
        "/oauth2/v2.0/devicecode",
        data={"client_id": cfg["outlook"]["graph_client_id"], "scope": scope or pd_lib.GRAPH_SCOPES},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def device_code_poll(cfg: dict, device_code: str, seconds: int = 150) -> dict | None:
    import requests
    tenant = cfg["outlook"].get("graph_tenant_id") or "organizations"
    deadline = time.time() + seconds
    interval = 5
    while time.time() < deadline:
        time.sleep(interval)
        r = requests.post(
            f"{pd_lib.LOGIN_BASE}/{tenant}/oauth2/v2.0/token",
            data={
                "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
                "client_id": cfg["outlook"]["graph_client_id"],
                "device_code": device_code,
            },
            timeout=30,
        )
        body = r.json()
        if r.status_code == 200:
            return body
        err = body.get("error")
        if err == "authorization_pending":
            continue
        if err == "slow_down":
            interval += 5
            continue
        raise RuntimeError(body.get("error_description", err))
    return None


def encode_sharing_url(url: str) -> str:
    """Turn a normal OneDrive/SharePoint 'Copy link' URL into the token the /shares
    endpoint needs: unpadded base64url, prefixed with 'u!'."""
    b64 = base64.b64encode(url.strip().encode("utf-8")).decode("ascii")
    return "u!" + b64.rstrip("=").replace("/", "_").replace("+", "-")


def resolve_onedrive_link(token: str, url: str) -> dict:
    """Resolve a sharing link to a driveItem. Returns id, name, driveId, webUrl."""
    import requests
    encoded = encode_sharing_url(url)
    r = requests.get(
        f"{pd_lib.GRAPH_BASE}/shares/{encoded}/driveItem",
        headers={"Authorization": f"Bearer {token}", "Prefer": "redeemSharingLink"},
        params={"$select": "id,name,webUrl,size,parentReference"},
        timeout=30,
    )
    if r.status_code == 404:
        raise RuntimeError("That link doesn't seem to point to a file — check it's the right one.")
    if r.status_code == 401:
        raise RuntimeError("Sign-in has expired. Click Connect to OneDrive again.")
    if r.status_code != 200:
        raise RuntimeError(f"Could not open that link ({r.status_code}): {r.text[:200]}")
    item = r.json()
    drive_id = (item.get("parentReference") or {}).get("driveId")
    if not drive_id or not item.get("id"):
        raise RuntimeError("Microsoft didn't return enough information to identify that file.")
    return {"drive_id": drive_id, "item_id": item["id"], "name": item.get("name", "spreadsheet.xlsx")}


def onedrive_download(token: str, drive_id: str, item_id: str) -> bytes:
    import requests
    r = requests.get(
        f"{pd_lib.GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    if r.status_code != 200:
        raise RuntimeError(f"Could not download the file ({r.status_code}): {r.text[:200]}")
    return r.content


def onedrive_upload(token: str, drive_id: str, item_id: str, data: bytes) -> None:
    import requests
    r = requests.put(
        f"{pd_lib.GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=data,
        timeout=60,
    )
    if r.status_code == 423:
        raise RuntimeError(
            "The file is currently open in Excel, so OneDrive has it locked. Close it there "
            "and click this button again."
        )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Could not save back to OneDrive ({r.status_code}): {r.text[:200]}")


def sidebar_onedrive(cfg: dict) -> None:
    """Connect-to-OneDrive UI. Independent token from the Outlook one, on purpose —
    see the ONEDRIVE_SCOPES comment above."""
    if not cfg["outlook"].get("graph_client_id"):
        st.caption("Not available yet — needs the same one-off setup as Outlook.")
        return
    if st.session_state.get("onedrive_token"):
        st.success("Connected to OneDrive")
        return

    if st.button("Connect to OneDrive", use_container_width=True):
        try:
            st.session_state["onedrive_device_code"] = device_code_start(cfg, ONEDRIVE_SCOPES)
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not start sign-in: {exc}")
        st.rerun()

    dc = st.session_state.get("onedrive_device_code")
    if dc:
        st.markdown(f"**Code:** `{dc['user_code']}`")
        st.markdown(f"1. Open {dc['verification_uri']}\n2. Enter the code\n3. Sign in")
        with st.spinner("Waiting for sign-in..."):
            try:
                tok = device_code_poll(cfg, dc["device_code"])
            except Exception as exc:  # noqa: BLE001
                st.error(
                    f"Sign-in failed: {exc}\n\nIf this mentions the scope or permission not "
                    "being found, the OneDrive permission hasn't been added to the app "
                    "registration yet — ask Mo."
                )
                tok = None
        if tok:
            st.session_state["onedrive_token"] = tok["access_token"]
            st.session_state.pop("onedrive_device_code", None)
            st.rerun()
        else:
            st.warning("Timed out. Click Connect again.")


def sidebar_outlook(cfg: dict) -> None:
    """Connect-to-Outlook UI. Lives in the sidebar so both the reply check and the
    draft creation can use the same session."""
    if not cfg["outlook"].get("graph_client_id"):
        st.info("Outlook isn't connected yet — you'll use the one-at-a-time route.")
        return
    if st.session_state.get("graph_token"):
        st.success("Connected to Outlook")
        return

    if st.button("Connect to Outlook", use_container_width=True):
        try:
            st.session_state["device_code"] = device_code_start(cfg)
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not start sign-in: {exc}")
        st.rerun()

    dc = st.session_state.get("device_code")
    if dc:
        st.markdown(f"**Code:** `{dc['user_code']}`")
        st.markdown(f"1. Open {dc['verification_uri']}\n2. Enter the code\n3. Sign in")
        with st.spinner("Waiting for sign-in..."):
            try:
                tok = device_code_poll(cfg, dc["device_code"])
            except Exception as exc:  # noqa: BLE001
                st.error(f"Sign-in failed: {exc}")
                tok = None
        if tok:
            st.session_state["graph_token"] = tok["access_token"]
            st.session_state.pop("device_code", None)
            st.rerun()
        else:
            st.warning("Timed out. Click Connect again.")


# ---------------------------------------------------------------------------
# Batch Email
# ---------------------------------------------------------------------------

EXAMPLE_TEMPLATE = """Dear {{First Name}},

Your membership reference is {{Reference Number}}. I am writing to confirm that
{{Company}} is now listed in this year's [member directory](https://example.com/directory).

If anything above needs correcting, reply to this email and I will see to it.

Kind regards,
Tooka"""


def batch_sidebar(cfg: dict) -> None:
    """Batch Email needs less in the sidebar: no AI key, and nothing to write back.

    Kept separate from the sequence workflows' sidebar rather than hiding pieces of that
    one, so neither has to reason about the other's state.
    """
    st.subheader("Your sign-off")
    st.text_area(
        "Email signature", value=cfg["sender"].get("signature_html", ""),
        key="sender_signature_html", height=100,
        help="Added under every draft, unless you tell the app your email already ends "
             "with its own sign-off.",
    )
    st.divider()
    st.subheader("Outlook")
    sidebar_outlook(cfg)


def batch_template_step() -> tuple[str, str, bool]:
    """Step 1 — the email itself. Returns the body, the subject, and whether to sign off."""
    st.header("1. The email you're sending")
    st.caption(
        "Write it once, exactly as you want it sent. Wherever a word changes per person, "
        "put the spreadsheet's column name in double braces: `{{First Name}}`, "
        "`{{Reference Number}}`. Everything else goes out untouched — no AI rewrites this."
    )

    with st.expander("Show me an example"):
        st.code(EXAMPLE_TEMPLATE, language=None)
        st.caption(
            "Capitals and spacing don't have to match your spreadsheet exactly — "
            "`{{first name}}` finds a column called “First Name”.\n\n"
            "A blank cell stops that row being drafted, because an email that says “Dear ,” "
            "is worse than one that never went. Where a bit genuinely doesn't apply to "
            "everyone, put a question mark on the end — `{{Job Title?}}` — and a blank cell "
            "is allowed: it comes out as nothing and the line closes up."
        )

    with st.expander("Putting a link in the email"):
        st.markdown(
            "This box holds plain text, and plain text has nowhere to keep a link's "
            "address. So a hyperlink copied out of Word or Outlook arrives here as just its "
            "words, with the address gone. Write it down instead, either way round:\n\n"
            "```\nHave a look at our [member directory](https://example.com/directory).\n"
            "\nOr just the address on its own: https://example.com/directory\n```\n\n"
            "Both become a proper clickable link in the finished email. **Uploading a Word "
            "file keeps its hyperlinks** and writes them in the same form, so you can see "
            "and edit exactly where each one points."
        )

    how = st.radio(
        "Where's the email?", ["Type or paste it", "Upload a Word file"],
        horizontal=True, label_visibility="collapsed",
    )

    if how == "Upload a Word file":
        up = st.file_uploader(
            "Upload the email", type=["docx", "txt", "md"],
            help="A .docx saved out of Word, or a plain .txt / .md file.",
        )
        if up is not None:
            # Only reseed when the file actually changes, or every rerun would wipe out the
            # corrections made in the box below.
            sig = f"{up.name}:{up.size}"
            if st.session_state.get("be_upload_sig") != sig:
                try:
                    st.session_state["be_body"] = be_lib.read_template_upload(
                        up.name, up.getvalue()
                    )
                    st.session_state["be_upload_sig"] = sig
                except Exception as exc:  # noqa: BLE001
                    st.error(str(exc))
            if up.name.lower().endswith(".docx"):
                st.caption(
                    "⚠️ Word formatting is not carried over: bold, colour, fonts and images "
                    "are dropped and the wording comes through as plain paragraphs. "
                    "**Hyperlinks are kept**, written as `[the words](the address)` so you "
                    "can check where each one goes. Read it in the box below and fix "
                    "anything that moved."
                )
        elif not st.session_state.get("be_body"):
            st.info("Upload the email, or switch to “Type or paste it”.")

    body = st.text_area(
        "The email", key="be_body", height=320,
        placeholder="Dear {{First Name}},\n\n...",
        help="A blank line starts a new paragraph.",
    )

    subject = st.text_input(
        "Subject line", key="be_subject",
        placeholder="Membership confirmation — {{Reference Number}}",
        help="The subject takes placeholders too, from the same columns as the email.",
    )

    signoff_in_template = be_lib.looks_like_signoff(body)
    append_signature = not st.checkbox(
        "My email already ends with my sign-off",
        value=signoff_in_template,
        help="Ticked, the signature in the sidebar is left off, so nobody gets “Kind "
             "regards, Tooka” twice.",
    )
    return body, subject, append_signature


def batch_list_step(max_rows: int):
    """Step 2 — the list. Returns (path, be_lib.Table)."""
    st.header("2. Your list")
    up = st.file_uploader(
        "Upload the spreadsheet", type=["xlsx", "xlsm", "csv"],
        help="One row per recipient. The file is only read, never changed.",
    )
    if not up:
        st.info("Upload the list to carry on.")
        st.stop()
    path = save_upload(up)

    sheet = None
    if path.suffix.lower() != ".csv":
        try:
            names, active = be_lib.sheet_names(path)
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not open that workbook: {exc}")
            st.stop()
        default = names.index(active) if active in names else 0
        sheet = st.selectbox(
            "Which sheet?", options=names, index=default,
            format_func=lambda n: f"{n}  ·  active in Excel" if n == active else n,
        )

    header_row = int(st.number_input(
        "Which row has the column names?", min_value=1, max_value=50, value=1,
        help="Usually row 1. Change it if your sheet starts with a title row.",
    ))

    try:
        table = be_lib.read_table(path, sheet, header_row)
    except Exception as exc:  # noqa: BLE001
        st.error(str(exc))
        st.stop()

    if not table.rows:
        st.error(
            f"No rows found under row {header_row}. Check you picked the right sheet and "
            "the right header row."
        )
        st.stop()

    st.write(f"**{len(table.rows)}** rows, **{len(table.headers)}** columns.")
    if table.duplicate_headers:
        st.warning(
            "Two columns share a name: "
            + ", ".join(f"“{h}”" for h in table.duplicate_headers)
            + ". Only the first of each is used — rename the other if you meant it."
        )
    if max_rows and len(table.rows) > max_rows:
        st.error(
            f"This list has **{len(table.rows)}** rows and Batch Email is capped at "
            f"**{max_rows}** per batch. Split it into smaller files, or ask Mo to raise "
            "`max_rows` in the config."
        )
        st.stop()
    with st.expander("The columns it found"):
        st.write(", ".join(f"“{h}”" for h in table.headers))
    return path, table


def batch_mapping_step(table, placeholders: list[str]) -> tuple[dict[str, str], str, str]:
    """Step 3 — which column feeds which placeholder. Returns (mapping, email col, label col).

    The other two workflows read fixed field names out of config.json. This one cannot: a
    batch list's columns are whatever that batch happens to have, so the matching is done
    here, against the headings actually in the file, and shown to the user to confirm.
    """
    st.header("3. Which column is which")
    keys = table.keys
    saved: dict = st.session_state.setdefault("be_map", {})
    auto = be_lib.auto_mapping(placeholders, keys)

    def column_picker(label: str, current: str, help_text: str = "") -> str:
        options = [""] + keys
        index = options.index(current) if current in options else 0
        return st.selectbox(
            label, options=options, index=index,
            format_func=lambda k: table.header_for(k) if k else "— pick a column —",
            help=help_text or None,
        )

    st.markdown("**Where are the email addresses?**")
    email_col = column_picker(
        "Email column",
        saved.get("__email__") or be_lib.guess_email_column(keys),
        "The address each draft is sent to. This is the one column the app cannot guess "
        "its way around.",
    )
    saved["__email__"] = email_col
    if not email_col:
        st.warning("Pick the column holding the email addresses.")
        st.stop()

    if not placeholders:
        st.info(
            "This email has no `{{placeholders}}` in it, so every draft will be word for "
            "word identical. That is allowed — but check it is what you meant."
        )
        return {}, email_col, ""

    matched = {k: v for k, v in auto.items() if not saved.get(k)}
    unmatched = [p for p in placeholders if not (saved.get(be_lib.norm(p)) or auto.get(be_lib.norm(p)))]

    if matched:
        st.success(
            "Matched on its own: "
            + ", ".join(f"`{{{{{p}}}}}` → **{table.header_for(auto[be_lib.norm(p)])}**"
                        for p in placeholders if be_lib.norm(p) in matched)
        )
    if unmatched:
        st.warning(
            f"{len(unmatched)} of the changing bits have no obvious column. Pick one for "
            "each — nothing can be drafted until every one is matched."
        )

    mapping: dict[str, str] = {}
    with st.expander("Every match, and how to change one", expanded=bool(unmatched)):
        for name in placeholders:
            key = be_lib.norm(name)
            current = saved.get(key) or auto.get(key, "")
            chosen = column_picker(f"`{{{{{name}}}}}`", current)
            saved[key] = chosen
            if chosen:
                mapping[key] = chosen

    still_missing = [p for p in placeholders if not mapping.get(be_lib.norm(p))]
    if still_missing:
        st.error(
            "No column chosen for: "
            + ", ".join(f"`{{{{{p}}}}}`" for p in still_missing)
            + ". Either pick one, or take it out of the email."
        )
        st.stop()

    # A recognisable name to label each draft with on screen. Nice to have, not required:
    # the address does the job if there is nothing better.
    label_col = ""
    for candidate in ("first name", "full name", "name", "recipient name", "contact",
                      "company", "company name"):
        if candidate in keys:
            label_col = candidate
            break
    return mapping, email_col, label_col


def batch_check_step(drafts: list) -> list:
    """Step 4 — the gate. Returns the rows fit to send; blocked ones are named, not dropped."""
    st.header("4. Check the list")
    ready = [d for d in drafts if d.ok]
    blocked = [d for d in drafts if not d.ok]
    dupes = [d for d in ready if d.warnings]

    c1, c2, c3 = st.columns(3)
    c1.metric("Ready", len(ready))
    c2.metric("Blocked", len(blocked))
    c3.metric("Duplicate addresses", len(dupes))

    if blocked:
        st.error(
            f"**{len(blocked)} row(s) cannot be drafted** and are left out. Fix them in the "
            "spreadsheet and upload it again — an email with a gap where the name should be "
            "is worse than one that never went."
        )
        st.dataframe(
            [{"Sheet row": d.row, "To": d.to or "(none)", "What's wrong": "; ".join(d.problems)}
             for d in blocked],
            hide_index=True, use_container_width=True,
        )
    if dupes:
        st.warning(
            "The same address appears more than once — they would get two emails: "
            + "; ".join(f"row {d.row} ({d.to})" for d in dupes[:10])
            + ("..." if len(dupes) > 10 else "")
        )
    if not ready:
        st.stop()
    if not blocked:
        st.success(f"All {len(ready)} rows have everything they need.")
    return ready


def batch_preview_step(ready: list) -> None:
    """Step 5 — read one in full before committing to the batch."""
    st.header("5. Read one first")
    st.caption(
        "Flick through a few. This is what catches the things no check can — a column that "
        "reads oddly mid-sentence, a reference in the wrong format."
    )
    pick = st.selectbox(
        "Which one?", options=list(range(len(ready))),
        format_func=lambda i: f"Row {ready[i].row} — {ready[i].label}",
    )
    d = ready[pick]
    # No `key` on these two, deliberately. A keyed widget reads its value out of session
    # state in preference to `value=`, so the preview would stick on whichever row was
    # shown first and quietly stop following the picker.
    st.text_input("Subject", value=d.subject, disabled=True)
    st.text_area("Message", value=d.body, height=280, disabled=True)
    st.caption(f"Goes to **{d.to}**")


def batch_email_flow(cfg: dict) -> None:
    """Batch Email, start to finish.

    A deliberately separate path from the sequence workflows rather than more branches
    inside main(): almost nothing is shared beyond getting the finished drafts into Outlook,
    and keeping them apart means a change here cannot put a live cold-call batch wrong.
    """
    max_rows = int(cfg["workflow"].get("max_rows") or 0)

    body, subject, append_signature = batch_template_step()
    if not body.strip():
        st.info("Write or upload the email to carry on.")
        st.stop()
    if not subject.strip():
        st.warning("Every email needs a subject line.")
        st.stop()

    placeholders = be_lib.find_placeholders(subject, body)
    if placeholders:
        st.caption(
            f"**{len(placeholders)} changing bit(s):** "
            + ", ".join(f"`{{{{{p}}}}}`" for p in placeholders)
        )

    # Shown whether or not any were found. A silent absence is exactly what someone who
    # pasted a hyperlink out of Outlook needs to see, because their link is the thing that
    # went missing on the way into a plain-text box.
    links = be_lib.find_links(body)
    if links:
        st.caption(
            f"**{len(links)} link(s):** "
            + ", ".join(
                (f"“{label}” → {url}" if label != url else url) for label, url in links[:6]
            )
            + ("..." if len(links) > 6 else "")
        )
    else:
        st.caption(
            "**No links found.** If your email is meant to have one, see *Putting a link in "
            "the email* above — pasting into a plain box loses the address."
        )

    path, table = batch_list_step(max_rows)
    mapping, email_col, label_col = batch_mapping_step(table, placeholders)

    drafts = be_lib.build_drafts(table, subject, body, mapping, email_col, label_col)
    ready = batch_check_step(drafts)
    batch_preview_step(ready)

    # ---- Step 6: make them --------------------------------------------------
    st.header("6. Make the drafts")
    how_many = int(st.number_input(
        "How many?", min_value=1, max_value=len(ready), value=len(ready),
        help="All of them, unless you want to try a handful in Outlook first.",
    ))

    if st.button(f"Make {how_many} drafts", type="primary", use_container_width=True):
        st.session_state["be_drafts"] = [
            {
                "row": d.row, "to": d.to, "name": d.label, "company": "",
                "subject": d.subject, "body": d.body,
                "warnings": list(d.warnings), "approved": True,
            }
            for d in ready[:how_many]
        ]
        st.session_state["be_log"] = {}

    made = st.session_state.get("be_drafts")
    if not made:
        st.stop()

    # ---- Step 7: read them --------------------------------------------------
    st.header("7. Read them")
    st.caption(
        "The wording is the same in all of them, so what is worth checking is the merged "
        "bits and who is on the list."
    )
    st.dataframe(
        [{"Sheet row": d["row"], "To": d["to"], "Subject": d["subject"]} for d in made],
        hide_index=True, use_container_width=True,
    )

    left_out = st.multiselect(
        "Leave anyone out?",
        options=[d["row"] for d in made],
        format_func=lambda r: next(
            f"Row {r} — {d['name']} ({d['to']})" for d in made if d["row"] == r
        ),
        help="Everyone is included unless you name them here.",
    )
    for d in made:
        d["approved"] = d["row"] not in set(left_out)

    with st.expander("Change one of them by hand"):
        pick = st.selectbox(
            "Which one?", options=list(range(len(made))),
            format_func=lambda i: f"Row {made[i]['row']} — {made[i]['name']}",
            key="be_edit_pick",
        )
        made[pick]["subject"] = st.text_input(
            "Subject", value=made[pick]["subject"], key=f"be_es{made[pick]['row']}"
        )
        made[pick]["body"] = st.text_area(
            "Message", value=made[pick]["body"], height=260, key=f"be_eb{made[pick]['row']}"
        )
        st.caption(
            "Edits stay on this one draft. Change the email itself in step 1 to change "
            "them all."
        )

    approved = [d for d in made if d["approved"]]
    st.info(f"**{len(approved)} of {len(made)}** ready to go into Outlook.")
    if not approved:
        st.stop()

    # ---- Step 8: into Outlook ----------------------------------------------
    st.header("8. Put them in Outlook")
    typed = (st.session_state.get("sender_signature_html") or "").strip()
    cfg["sender"]["signature_html"] = typed if append_signature else ""
    if not append_signature:
        st.caption("Your email ends with its own sign-off, so the sidebar signature is left off.")

    def record(done: list[dict], failed: list[tuple]) -> None:
        log = st.session_state.setdefault("be_log", {})
        for d in done:
            log[d["row"]] = ("Draft created in Outlook", "")
        for d, err in failed:
            log[d["row"]] = ("Failed", err)

    deliver_drafts(approved, cfg, f"batch_{date.today():%Y-%m-%d}", on_pushed=record)

    # ---- The log -----------------------------------------------------------
    st.divider()
    st.subheader("Keep a record")
    st.caption(
        "One line per row of your list, including the blocked ones. Worth keeping for its "
        "own sake, and next week it is the only way to tell who already had theirs."
    )
    log = st.session_state.get("be_log") or {}
    approved_rows = {d["row"] for d in approved}
    rows = []
    for d in drafts:
        outcome, detail = log.get(d.row, ("", ""))
        if not outcome:
            if not d.ok:
                outcome, detail = "Blocked", "; ".join(d.problems)
            elif d.row not in approved_rows:
                outcome, detail = "Left out", ""
            else:
                outcome, detail = "Drafted, not yet in Outlook", ""
        rows.append({"row": d.row, "to": d.to, "subject": d.subject,
                     "outcome": outcome, "detail": detail})
    st.download_button(
        "Download the log (CSV)", data=be_lib.log_csv(rows),
        file_name=f"batch_log_{date.today():%Y-%m-%d}.csv", mime="text/csv",
        use_container_width=True,
    )


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

def main() -> None:
    st.title("✉️ Prospect Drafter")

    # ---- Which workflow ------------------------------------------------------
    # Chosen before anything else, including the AI key check: Batch Email has no AI in it
    # at all, and a red "AI key problem" banner over a workflow that never calls the AI is
    # just a lie the user has to learn to ignore.
    base = get_config()
    names = pd_lib.workflow_names(base)

    if len(names) > 1:
        workflow = st.radio(
            "Workflow",
            options=names,
            format_func=lambda n: pd_lib.workflow_label(base, n),
            horizontal=True,
            key="workflow",
        )
    else:
        workflow = names[0] if names else None

    cfg = get_config(workflow)
    wf = cfg.get("workflow", {})
    mode = str(wf.get("mode") or "sequence").lower()

    # Drafts belong to the workflow they were written for. Switching workflow clears them so
    # a batch can never be written back through the other workflow's settings. This matters
    # more now that both sequence workflows patch the same workbook: they use different
    # sheets and different Status columns (C vs A), so a mismatched write-back would put a
    # status into the wrong column of the wrong sheet. Batch Email's own working state goes
    # the same way, so a half-matched batch cannot leak into the next workflow.
    if st.session_state.get("drafted_workflow") not in (None, workflow):
        for key in ("drafts", "stage", "source_path", "replied",
                    "be_drafts", "be_log", "be_map", "be_upload_sig"):
            st.session_state.pop(key, None)
        st.session_state.pop("drafted_workflow", None)

    st.caption(
        "One email you have already written, merged with a list. **Nothing is ever sent** — "
        "you read the drafts, then they go into Outlook for you to send yourself."
        if mode == "batch" else
        "Reads each prospect's website and writes a personalised email. **Nothing is ever "
        "sent** — you read every draft, then it goes into Outlook for you to send yourself."
    )
    if wf.get("blurb"):
        st.caption(wf["blurb"])

    # ---- Batch Email takes its own path ------------------------------------
    if mode == "batch":
        with st.sidebar:
            batch_sidebar(cfg)
        st.session_state["drafted_workflow"] = workflow
        batch_email_flow(cfg)
        return

    # ---- AI key check -------------------------------------------------------
    # A live check, not a guess from the key's shape: this app once flagged a real, working
    # key as broken purely because of its prefix. Only asking Google settles it, and it's
    # worth doing before anything else because a broken key means every draft below silently
    # degrades to the plain [FILL THIS IN] template - better the user hears that now than
    # discovers it three drafts in.
    ai_key = secret(base["ai"]["api_key_env"])
    key_check = cached_ai_key_check(ai_key, base["ai"]["model"], base["ai"].get("base_url"))
    if key_check["ok"] is False:
        st.error(
            f"**AI key problem** — {key_check['message']}\n\n"
            "The app still works without it: every draft will fall back to the plain "
            "template with `[FILL THIS IN]` gaps for you to complete by hand, instead of a "
            "written email."
        )
    elif key_check["ok"] is None:
        st.warning(f"Could not confirm the AI key is working — {key_check['message']}")

    # Branch on what a workflow *does*, not on which one it is. Both sequence workflows are
    # status-driven against the same workbook, and the only differences left between them
    # live in config: which sheet, which columns, whether there is a score to filter on.
    seq_gate = str(cfg["sequence"].get("gate", "touches")).lower()
    uses_status = seq_gate == "status"
    has_fit_score = bool(wf.get("fit_score"))
    patch_writeback = str((wf.get("writeback") or {}).get("mode", "")).lower() == "patch"

    seq = cfg["sequence"]
    labels = {int(k): v for k, v in seq["labels"].items()}
    wait_days = int(seq.get("wait_days", 7))

    with st.sidebar:
        st.subheader("Your details")
        st.text_input("Your name", value=cfg["sender"]["your_name"], key="sender_your_name")
        st.text_input("Your company", value=cfg["sender"]["your_company"],
                      key="sender_your_company")
        st.text_area("Email signature", value=cfg["sender"].get("signature_html", ""),
                     key="sender_signature_html", height=100,
                     help="Added to the bottom of every draft.")
        st.divider()
        st.subheader("Outlook")
        sidebar_outlook(cfg)
        st.divider()
        st.subheader("OneDrive")
        st.caption("Optional — lets you load and save the spreadsheet without downloading it.")
        sidebar_onedrive(cfg)
        st.divider()
        if key_check["ok"] is True:
            st.caption("AI key active ✓")
        elif key_check["ok"] is False:
            st.error(key_check["message"])
        else:
            st.caption(f"AI key configured, unverified — {key_check['message']}")

    # ---- Step 1: the list ---------------------------------------------------
    st.header("1. Your prospect list")

    source_choice = st.radio(
        "Where's your spreadsheet?", ["Upload a file", "Load from OneDrive"],
        horizontal=True, label_visibility="collapsed",
    )

    path = None

    if source_choice == "Upload a file":
        st.session_state.pop("onedrive_source", None)
        uploaded = st.file_uploader(
            "Upload your spreadsheet", type=["xlsx", "xlsm", "csv"],
            help="The file you keep your prospects in. It is never modified.",
        )
        if not uploaded:
            st.info("Upload a spreadsheet to begin.")
            st.stop()
        path = save_upload(uploaded)

    else:
        if not st.session_state.get("onedrive_token"):
            st.info("Connect to OneDrive in the sidebar first.")
            st.stop()

        link = st.text_input(
            "Paste the OneDrive link to your spreadsheet",
            help="In OneDrive or Excel Online: right-click the file → Copy link.",
        )
        loaded = st.session_state.get("onedrive_source")

        if st.button("Load from OneDrive", use_container_width=True, disabled=not link.strip()):
            try:
                info = resolve_onedrive_link(st.session_state["onedrive_token"], link.strip())
                data = onedrive_download(
                    st.session_state["onedrive_token"], info["drive_id"], info["item_id"]
                )
                dest = APP_DIR / "_uploads"
                dest.mkdir(exist_ok=True)
                local_path = dest / info["name"]
                local_path.write_bytes(data)
                st.session_state["onedrive_source"] = {**info, "local_path": str(local_path)}
                st.success(f"Loaded **{info['name']}**.")
                st.rerun()
            except Exception as exc:  # noqa: BLE001
                st.error(str(exc))

        if not loaded:
            st.stop()
        st.caption(f"Using **{loaded['name']}** from OneDrive.")
        path = Path(loaded["local_path"])

    cfg["spreadsheet"]["path"] = str(path)

    chosen_sheet = sheet_selector(path, cfg)
    if chosen_sheet:
        # Everything downstream reads this, including the Cold Call write-back, so the
        # sheet that gets patched is always the sheet that was read.
        cfg["spreadsheet"]["sheet_name"] = chosen_sheet

    try:
        prospects = pd_lib.read_prospects(cfg)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not read that file: {exc}")
        st.stop()

    st.write(f"**{len(prospects)}** rows read.")

    # ---- Pick the fit scores to work, where the sheet has them ---------------
    if has_fit_score:
        prospects = fit_score_filter(prospects, cfg)

    # The number of stages is a config choice, not a fixed 3 - Cold Call now runs a fourth,
    # final follow-up that Internal lead doesn't have. Everything below reads it from
    # `labels` rather than assuming a count, so a workflow can add or drop a stage in
    # config.json alone.
    stages = sorted(labels)

    today = date.today()
    buckets: dict[int, list] = {}
    reasons: dict[int, list[tuple]] = {}
    for stage in stages:
        ok, no = [], []
        for p in prospects:
            good, why = pd_lib.eligible_for_stage(p, stage, cfg, today)
            (ok if good else no).append(p if good else (p, why))
        buckets[stage] = ok
        reasons[stage] = no

    for col, stage in zip(st.columns(len(stages)), stages):
        col.metric(labels[stage], len(buckets[stage]))

    # ---- Step 2: which email ------------------------------------------------
    st.header("2. Which email are you sending?")

    stage = st.radio(
        "Stage",
        options=stages,
        format_func=lambda s: f"{labels[s]} — {len(buckets[s])} ready",
        horizontal=True,
        label_visibility="collapsed",
    )

    if uses_status:
        flow = (seq.get("status_flow") or {}).get(str(stage), {})
        needs = ", ".join(f"`{x}`" if x else "blank" for x in flow.get("from", [])) or "—"
        sets = pd_lib.next_status(cfg, stage)
        st.caption(
            f"Rows whose **Status** is {needs}"
            + (f", last contacted at least {wait_days} days ago" if stage > 1 else "")
            + (f". Once drafted, their Status becomes `{sets}`." if sets else ".")
        )
    elif stage == 1:
        st.caption("People who haven't been emailed yet.")
    else:
        st.caption(
            f"People who received email {stage - 1}, haven't replied, and were last "
            f"contacted at least {wait_days} days ago."
        )

    ready = buckets[stage]

    if not ready:
        st.warning(f"Nobody is ready for the {labels[stage].lower()} right now.")
        with st.expander("Why not?"):
            for p, why in reasons[stage][:40]:
                st.write(f"- **{p.company or p.email}** — {why}")
        st.stop()

    # Optional mailbox check before drafting a follow-up
    if stage > 1:
        if st.session_state.get("graph_token"):
            if st.button("Check who has replied first", use_container_width=True):
                excluded = []
                bar = st.progress(0.0, text="Checking your mailbox...")
                for i, p in enumerate(list(ready), 1):
                    since = p.latest_contact or today
                    try:
                        if graph_has_reply(st.session_state["graph_token"], p.email, since):
                            excluded.append(p.email.lower())
                    except Exception as exc:  # noqa: BLE001
                        st.warning(f"Could not check {p.email}: {exc}")
                        break
                    bar.progress(i / len(ready))
                bar.empty()
                st.session_state["replied"] = excluded
                st.success(
                    f"{len(excluded)} of these have replied — removed from the list."
                    if excluded else "Nobody on this list has replied."
                )
        else:
            st.caption(
                "Tip: connect Outlook in the sidebar and the app can check who has already "
                "replied. Until then, mark them `Replied` in the Status column yourself."
            )

    replied = set(st.session_state.get("replied") or [])
    if replied:
        ready = [p for p in ready if p.email.lower() not in replied]
        st.info(f"{len(replied)} prospect(s) excluded because they replied.")
        if not ready:
            st.stop()

    no_site = [p for p in ready if not p.website]
    if no_site:
        st.caption(
            f"{len(no_site)} of these have no website. They'll get the generic version of "
            "the template — nothing will be invented about them."
        )

    with st.expander("Check the columns were read correctly"):
        p0 = ready[0]
        st.write({
            "First name": p0.first_name, "Last name": p0.last_name, "Company": p0.company,
            "Job title": p0.job_title or "(blank)", "Email": p0.email,
            "Website": p0.website or "(blank)",
            "Emails sent so far": p0.touches,
            "Last contact": str(p0.latest_contact) if p0.latest_contact else "(none)",
            "Extra context": p0.context or "(none)",
        })

    # ---- Step 3: draft ------------------------------------------------------
    st.header("3. Write the drafts")

    template_path = pd_lib.stage_template_path(cfg, stage, APP_DIR)
    if not template_path.exists():
        st.error(f"The template for this stage is missing ({template_path.name}). Tell Mo.")
        st.stop()

    how_many = st.number_input(
        "How many prospects?", min_value=1, max_value=len(ready),
        value=min(3, len(ready)),
        help="Start small so you can check the tone before doing the whole list.",
    )

    called_first = False
    if stage == 1:
        called_first = st.checkbox(
            "I tried phoning these people first",
            value=bool(
                wf.get("called_first_default") or cfg["sender"].get("mention_prior_call")
            ),
            help=(
                "The first email can mention that you tried to call. Only tick this if you "
                "actually did — otherwise the sentence is left out."
            ),
        )
        if called_first:
            st.caption(
                "The email will say you tried to call and couldn't get through. Make sure "
                "that's true for everyone in this batch."
            )

    if st.button(f"Write the {labels[stage].lower()} drafts", type="primary",
                 use_container_width=True):
        queue = ready[: int(how_many)]
        template = template_path.read_text("utf-8")
        system_prompt = pd_lib.build_system_prompt(cfg, template, stage, called_first)
        os.environ["GEMINI_API_KEY"] = secret("GEMINI_API_KEY")

        # Some stages send fixed, approved copy rather than an AI-adapted email — see
        # is_verbatim_stage(). There's nothing for research to feed into on those stages, so
        # skip the website fetch, and the "researched" flag would otherwise wrongly mark every
        # one of these as needing a look even though nothing went wrong.
        verbatim = pd_lib.is_verbatim_stage(cfg, stage)
        fixed_subject = cfg["sequence"].get("fixed_subject")

        drafts = []
        bar = st.progress(0.0, text="Starting...")
        for i, p in enumerate(queue, 1):
            bar.progress((i - 1) / len(queue), text=f"{p.full_name or p.email} — {p.company}")
            research, warns = ("", [])
            if p.website and not verbatim:
                research, warns = pd_lib.fetch_site_text(p.website, cfg)
            try:
                out = pd_lib.call_gemini(cfg, system_prompt, pd_lib.build_user_prompt(p, research))
                ai_ok = True
            except Exception as exc:  # noqa: BLE001
                warns.append(f"AI unavailable: {exc}")
                out = pd_lib.fallback_fill(template, p, cfg)
                ai_ok = False
            drafts.append({
                "to": p.email, "name": p.full_name or p.email, "company": p.company,
                "subject": fixed_subject or out["subject"],
                "body": html_to_plain(out["body_html"]),
                "note": out["personalisation_note"],
                "researched": True if verbatim else len(research) >= 200,
                "ai_ok": ai_ok, "warnings": warns, "approved": True,
                "touches_after": p.touches + 1,
                # For the Cold Call Status write-back.
                "row": p.row,
                "next_status": pd_lib.next_status(cfg, stage),
                "needs_first_date": p.first_contact is None,
            })
            if i < len(queue):
                time.sleep(float(cfg["ai"]["delay_seconds"]))
        bar.progress(1.0, text="Done")
        st.session_state["drafts"] = drafts
        st.session_state["stage"] = stage
        st.session_state["source_path"] = str(path)
        st.session_state["drafted_workflow"] = workflow

    drafts = st.session_state.get("drafts")
    if not drafts:
        st.stop()

    drafted_stage = st.session_state.get("stage", stage)
    if drafted_stage != stage:
        st.warning(
            f"The drafts below are the **{labels[drafted_stage].lower()}**. Click the button "
            "above to write drafts for the stage you just selected."
        )

    # The sign-off follows the stage where the workflow defines one, because approved copy does
    # not sign off the same way throughout a sequence. Anything typed into the sidebar wins.
    typed = (st.session_state.get("sender_signature_html") or "").strip()
    if not typed or typed == (load_config()["sender"].get("signature_html") or "").strip():
        cfg["sender"]["signature_html"] = pd_lib.stage_signature(cfg, drafted_stage)

    # ---- Step 4: review -----------------------------------------------------
    st.header("4. Read them")
    weak = [d for d in drafts if not d["researched"] or not d["ai_ok"]]
    if weak:
        st.warning(
            f"{len(weak)} draft(s) marked **needs a look** — the website couldn't be read, so "
            "they use the plain template. Read those first."
        )
    st.caption("Untick anything you don't want. Edit the wording freely — it's just text.")

    for i, d in enumerate(drafts):
        flag = "" if (d["researched"] and d["ai_ok"]) else "  ⚠️ needs a look"
        with st.expander(f"{d['name']} — {d['company']}{flag}", expanded=(i == 0)):
            d["approved"] = st.checkbox("Include this one", value=d["approved"], key=f"ap{i}")
            st.caption(f"To: {d['to']}  ·  Angle: {d['note'] or '—'}")
            d["subject"] = st.text_input("Subject", value=d["subject"], key=f"su{i}")
            d["body"] = st.text_area("Message", value=d["body"], height=220, key=f"bo{i}")
            if d["warnings"]:
                st.caption("Notes: " + "; ".join(d["warnings"]))

    approved = [d for d in drafts if d["approved"]]
    st.info(f"**{len(approved)} of {len(drafts)}** ready to go into Outlook.")
    if not approved:
        st.stop()

    # ---- Step 5: deliver ----------------------------------------------------
    st.header("5. Put them in Outlook")
    deliver_drafts(approved, cfg, f"drafts_stage{drafted_stage}")

    # ---- Bookkeeping --------------------------------------------------------
    st.divider()
    st.subheader("Keep your spreadsheet up to date")
    st.caption(
        "This records who you just handled and when, which is what makes the follow-ups "
        "work later."
    )
    onedrive_source = st.session_state.get("onedrive_source")
    source_path = Path(st.session_state["source_path"])

    if patch_writeback:
        next_stat = pd_lib.next_status(cfg, drafted_stage)
        st.write(
            f"Sets **Status** to `{next_stat}` and updates the contact dates for the "
            f"**{len(approved)}** rows below. Everything else in the workbook is left "
            "exactly as it is."
            if next_stat else
            f"Updates the contact dates for the **{len(approved)}** rows below."
        )
        with st.expander("Which rows change"):
            for d in approved:
                bits = [f"Status → `{next_stat}`"] if next_stat else []
                if d.get("needs_first_date"):
                    bits.append("First Contact Date set")
                bits.append(f"Last Contact Date → {today:%Y-%m-%d}")
                st.write(f"- Row {d['row']} · **{d['company'] or d['to']}** — " + ", ".join(bits))
        try:
            updated_bytes = status_writeback(source_path, approved, cfg, today)
        except Exception as exc:  # noqa: BLE001
            updated_bytes = None
            st.error(f"Could not prepare the update: {exc}")
    else:
        updates = {
            d["to"].lower(): {"touches": d["touches_after"], "date": today} for d in approved
        }
        try:
            updated_bytes = spreadsheet_with_progress(source_path, updates, cfg)
        except Exception as exc:  # noqa: BLE001
            updated_bytes = None
            st.caption(f"(Could not build the updated spreadsheet: {exc})")

    if updated_bytes is not None and onedrive_source:
        if st.button("Save back to OneDrive", type="primary", use_container_width=True):
            try:
                onedrive_upload(
                    st.session_state["onedrive_token"],
                    onedrive_source["drive_id"], onedrive_source["item_id"], updated_bytes,
                )
                st.success(f"Saved back to **{onedrive_source['name']}** in OneDrive.")
            except Exception as exc:  # noqa: BLE001
                st.error(str(exc))
        st.caption("Or keep a local copy too:")

    if updated_bytes is not None:
        stem = source_path.stem if patch_writeback else "prospects"
        st.download_button(
            "Download updated spreadsheet",
            data=updated_bytes,
            file_name=f"{stem}_updated_{today:%Y-%m-%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
