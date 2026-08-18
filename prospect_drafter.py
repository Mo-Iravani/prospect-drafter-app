#!/usr/bin/env python3
"""
prospect_drafter.py — research prospects and draft personalised emails into Outlook.

Two-phase by design:

    1. research  →  reads your spreadsheet, visits each prospect's website,
                    asks an AI to write the email from your template, and
                    writes everything into an editable review file.

    2. push      →  reads the review file and creates drafts in Outlook for
                    every entry you marked APPROVED. Nothing is ever sent.

Usage:
    python prospect_drafter.py init
    python prospect_drafter.py models
    python prospect_drafter.py research [--limit N] [--config config.json]
    python prospect_drafter.py push [--backend eml|graph|com] [--config config.json]
    python prospect_drafter.py auth          (Graph device-code sign-in)
    python prospect_drafter.py status

Licence: yours. No warranty. Review every draft before sending.
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
import time
import urllib.parse
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

try:
    import requests
except ImportError:
    sys.exit("Missing dependency. Run:  pip install requests openpyxl beautifulsoup4 lxml")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GEMINI_BASE = "https://generativelanguage.googleapis.com/v1beta"
GRAPH_BASE = "https://graph.microsoft.com/v1.0"
LOGIN_BASE = "https://login.microsoftonline.com"
GRAPH_SCOPES = "offline_access Mail.ReadWrite User.Read"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

ABOUT_HINTS = (
    "about", "about-us", "aboutus", "who-we-are", "our-story",
    "services", "what-we-do", "solutions", "company",
)

REVIEW_SEPARATOR = "=" * 78

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULT_CONFIG: dict[str, Any] = {
    "spreadsheet": {
        "path": "prospects.xlsx",
        "sheet_name": None,
        "header_row": 1,
        "columns": {
            "company": "Company",
            "first_name": "Contact First Name",
            "last_name": "Contact Last Name",
            "job_title": "Job Title",
            "email": "Email",
            "website": "Website",
            "linkedin": "LinkedIn URL",
            "notes": "Notes",
            "status": "Status",
            # Cold Call only: the score column the user filters on.
            "fit_score": "",
            # Owned by the tool, not by you. Created automatically if absent.
            "touches": "Touches",
            "first_contact_date": "First Contact Date",
            "last_contact_date": "Last Contact Date",
        },
        # Any other columns you want the AI to see as context, by exact header name.
        "context_columns": [],
        # A Status starting with any of these means: stop, never contact again.
        "stop_statuses": [
            "replied", "responded", "skip", "do not contact", "unsubscribed",
            "bounced", "closed", "won", "lost", "customer",
        ],
        # Kept for the single-shot CLI: statuses that mean "already handled".
        "skip_when_status_in": ["drafted", "sent", "skip", "do not contact"],
        # {exact column header: [values that mean "never contact this row"]}. The column
        # must also be listed in context_columns so it gets read.
        "skip_when": {},
        # Phrases that mean "stop", matched anywhere in Status rather than at the start.
        # For sheets whose Status holds free-text judgement notes.
        "stop_contains": [],
        # {"1": [phrases meaning one email has gone], "2": [...]}, matched anywhere in
        # Status. Only consulted when there is no Touches column value on the row.
        "touch_patterns": {},
    },
    "sequence": {
        # How the app decides which email a prospect is due.
        #   "touches" - count the emails already sent, from the Touches column (Internal lead)
        #   "status"  - read the stage straight off the Status column (Cold Call)
        "gate": "touches",
        # Days that must pass since the last contact before a follow-up is offered.
        "wait_days": 7,
        # Stage number -> template file. Stage 1 is the first email.
        "templates": {
            "1": "template.md",
            "2": "template_followup_1.md",
            "3": "template_followup_2.md",
        },
        "labels": {
            "1": "First email",
            "2": "First follow-up",
            "3": "Second follow-up",
        },
    },
    "template_path": "template.md",
    "sender": {
        "your_name": "Mo Iravani",
        "your_title": "",
        "your_company": "Gateway Global LTD",
        "what_you_do": "One or two sentences on what Gateway Global does and who it helps.",
        "signature_html": "<p>Best regards,<br>Mo Iravani<br>Gateway Global LTD</p>",
    },
    "research": {
        "enabled": True,
        "timeout_seconds": 12,
        "max_chars": 6000,
        "follow_about_page": True,
    },
    "ai": {
        "provider": "gemini",
        "model": "gemini-flash-lite-latest",
        "api_key_env": "GEMINI_API_KEY",
        "delay_seconds": 3,
        "max_words": 130,
        "language": "British English",
        "base_url": None,
        "extra_rules": [
            "Never invent facts about the prospect or their company.",
            "No flattery openers such as 'I was blown away by'.",
            "One specific, verifiable observation about their business.",
            "One clear ask at the end.",
            "Do not use em-dashes.",
        ],
    },
    "output": {
        "review_file": "review/drafts_review.md",
        "eml_dir": "review/eml",
        "state_file": "review/state.json",
    },
    "outlook": {
        "backend": "eml",
        "graph_client_id": "",
        "graph_tenant_id": "organizations",
        "token_cache": "review/.token_cache.json",
    },
}


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        sys.exit(f"No config found at {path}. Run:  python prospect_drafter.py init")
    cfg = json.loads(path.read_text(encoding="utf-8"))
    return deep_merge(DEFAULT_CONFIG, cfg)


def deep_merge(base: dict, override: dict) -> dict:
    out = dict(base)
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = deep_merge(out[k], v)
        else:
            out[k] = v
    return out


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Prospect:
    row: int
    company: str = ""
    first_name: str = ""
    last_name: str = ""
    job_title: str = ""
    email: str = ""
    website: str = ""
    linkedin: str = ""
    notes: str = ""
    status: str = ""
    fit_score: str = ""
    context: dict[str, str] = field(default_factory=dict)
    touches: int = 0
    first_contact: date | None = None
    last_contact: date | None = None

    @property
    def latest_contact(self) -> date | None:
        return self.last_contact or self.first_contact

    @property
    def full_name(self) -> str:
        return " ".join(x for x in (self.first_name, self.last_name) if x).strip()

    @property
    def key(self) -> str:
        return (self.email or f"{self.company}#{self.row}").strip().lower()


@dataclass
class DraftResult:
    prospect: Prospect
    subject: str = ""
    body_html: str = ""
    personalisation_note: str = ""
    research_ok: bool = False
    ai_ok: bool = False
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Spreadsheet
# ---------------------------------------------------------------------------

def read_prospects(cfg: dict) -> list[Prospect]:
    from openpyxl import load_workbook

    sc = cfg["spreadsheet"]
    path = Path(sc["path"]).expanduser()
    if not path.exists():
        sys.exit(f"Spreadsheet not found: {path}")

    if path.suffix.lower() == ".csv":
        return _read_csv(path, sc)

    wb = load_workbook(path, data_only=True, read_only=True)
    # Never a bare wb[name]: a configured sheet that is missing from this particular file
    # would raise KeyError, which is a crash rather than an explanation.
    ws = wb[resolve_sheet_name(list(wb.sheetnames), sc.get("sheet_name"))]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Spreadsheet is empty.")

    hdr_idx = int(sc.get("header_row", 1)) - 1
    headers = [str(h).strip() if h is not None else "" for h in rows[hdr_idx]]
    lookup = {h.lower(): i for i, h in enumerate(headers) if h}

    prospects: list[Prospect] = []
    for offset, raw in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        prospects.append(
            _build_prospect(raw, lookup, sc["columns"], offset, sc.get("context_columns", []),
                            sc.get("touch_patterns"))
        )
    return prospects


def _read_csv(path: Path, sc: dict) -> list[Prospect]:
    import csv
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = list(csv.reader(fh))
    hdr_idx = int(sc.get("header_row", 1)) - 1
    headers = [h.strip() for h in reader[hdr_idx]]
    lookup = {h.lower(): i for i, h in enumerate(headers) if h}
    out = []
    for offset, raw in enumerate(reader[hdr_idx + 1:], start=hdr_idx + 2):
        if not any(str(c).strip() for c in raw):
            continue
        out.append(
            _build_prospect(raw, lookup, sc["columns"], offset, sc.get("context_columns", []),
                            sc.get("touch_patterns"))
        )
    return out


def _build_prospect(
    raw, lookup: dict, colmap: dict, row_no: int, context_columns: list[str] | None = None,
    touch_patterns: dict | None = None,
) -> Prospect:
    def by_header(header: str) -> str:
        header = str(header or "").strip().lower()
        if not header or header not in lookup:
            return ""
        idx = lookup[header]
        if idx >= len(raw):
            return ""
        v = raw[idx]
        return "" if v is None else str(v).strip()

    def val(field_name: str) -> str:
        return by_header(colmap.get(field_name))

    context = {}
    for header in context_columns or []:
        v = by_header(header)
        if v:
            context[str(header).strip()] = v

    status = val("status")
    touches = infer_touches(status, by_header(colmap.get("touches")), touch_patterns)

    return Prospect(
        context=context,
        touches=touches,
        first_contact=parse_date(by_header(colmap.get("first_contact_date"))),
        last_contact=parse_date(by_header(colmap.get("last_contact_date"))),
        row=row_no,
        company=val("company"),
        first_name=val("first_name"),
        last_name=val("last_name"),
        job_title=val("job_title"),
        email=val("email"),
        website=val("website"),
        linkedin=val("linkedin"),
        notes=val("notes"),
        status=status,
        fit_score=normalise_score(val("fit_score")),
    )


# ---------------------------------------------------------------------------
# Website research
# ---------------------------------------------------------------------------

def normalise_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if not re.match(r"^https?://", url, re.I):
        url = "https://" + url.lstrip("/")
    return url


def fetch_site_text(url: str, cfg: dict) -> tuple[str, list[str]]:
    """Return (text, warnings). Never raises."""
    warnings: list[str] = []
    url = normalise_url(url)
    if not url:
        return "", ["No website in the spreadsheet."]

    rc = cfg["research"]
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "en-GB,en;q=0.9"})

    pages: list[str] = []
    home_html = _get(session, url, rc["timeout_seconds"], warnings)
    if home_html is None:
        return "", warnings

    pages.append(_html_to_text(home_html))

    if rc.get("follow_about_page", True):
        sub = _find_about_link(home_html, url)
        if sub:
            sub_html = _get(session, sub, rc["timeout_seconds"], warnings)
            if sub_html:
                pages.append(_html_to_text(sub_html))

    text = "\n\n".join(p for p in pages if p).strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    if len(text) > rc["max_chars"]:
        text = text[: rc["max_chars"]] + "\n[truncated]"
    if len(text) < 200:
        warnings.append("Very little usable text found on the website.")
    return text, warnings


def _get(session, url: str, timeout: int, warnings: list[str]) -> str | None:
    try:
        r = session.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code >= 400:
            warnings.append(f"{url} returned HTTP {r.status_code}.")
            return None
        ctype = r.headers.get("Content-Type", "")
        if "html" not in ctype and "xml" not in ctype:
            warnings.append(f"{url} is not an HTML page ({ctype or 'unknown type'}).")
            return None
        return r.text
    except requests.RequestException as exc:
        warnings.append(f"Could not reach {url}: {type(exc).__name__}.")
        return None


def _find_about_link(html: str, base_url: str) -> str | None:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    base_host = urllib.parse.urlparse(base_url).netloc
    for a in soup.find_all("a", href=True):
        href = a["href"]
        absolute = urllib.parse.urljoin(base_url, href)
        parsed = urllib.parse.urlparse(absolute)
        if parsed.netloc and parsed.netloc != base_host:
            continue
        slug = (parsed.path or "").strip("/").lower()
        if not slug:
            continue
        last = slug.split("/")[-1]
        if any(h == last or h in last for h in ABOUT_HINTS):
            return absolute
    return None


def _html_to_text(html: str) -> str:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "svg", "nav", "footer", "form", "iframe"]):
        tag.decompose()
    parts: list[str] = []
    title = soup.find("title")
    if title and title.get_text(strip=True):
        parts.append("PAGE TITLE: " + title.get_text(strip=True))
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        parts.append("META DESCRIPTION: " + meta["content"].strip())
    body_text = soup.get_text("\n", strip=True)
    parts.append(body_text)
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# AI drafting
# ---------------------------------------------------------------------------

STAGE_GUIDANCE = {
    1: "This is the FIRST email to this person.",
    2: (
        "This is a FOLLOW-UP. They received your first email and did not reply.\n"
        "- Do not repeat the first email. Lead with something new: a different angle, a\n"
        "  useful observation, or a lighter ask.\n"
        "- Do not guilt them, and never write 'just bumping this' or 'did you see my email'.\n"
        "- Keep it shorter than the first email. Under 80 words is ideal.\n"
        "- Assume they are busy, not uninterested."
    ),
    3: (
        "This is the SECOND and FINAL follow-up. Two emails have gone unanswered.\n"
        "- Be brief and gracious. Four sentences at most.\n"
        "- Make it easy to say no, or to say 'not now'. Offer to stop following up.\n"
        "- No new pitch, no pressure, no third restatement of the offer.\n"
        "- This email should leave the door open, not push on it."
    ),
}


def build_system_prompt(
    cfg: dict, template: str, stage: int = 1, called_first: bool = False
) -> str:
    s = cfg["sender"]
    ai = cfg["ai"]
    rules = "\n".join(f"- {r}" for r in ai.get("extra_rules", []))
    stage_note = STAGE_GUIDANCE.get(stage, STAGE_GUIDANCE[1])
    if called_first:
        stage_note += (
            "\n\nA PHONE CALL WAS ATTEMPTED before this email. Where the template provides a"
            " sentence about having tried to call, include it."
        )
    else:
        stage_note += (
            "\n\nNO PHONE CALL WAS MADE. If the template contains a sentence about having tried"
            " to call, OMIT IT COMPLETELY. Do not state or imply that anyone tried to phone"
            " this person. This is a factual claim and it would be untrue."
        )
    return f"""{stage_note}

You write short, credible B2B outreach emails on behalf of {s['your_name']}\
{(', ' + s['your_title']) if s.get('your_title') else ''} at {s['your_company']}.

About the sender's business:
{s['what_you_do']}

You will be given a prospect's details and text scraped from their website. Write ONE email.

Hard rules:
- Write in {ai['language']}.
- Maximum {ai['max_words']} words in the body, excluding the greeting and sign-off.
- Follow the structure and intent of the TEMPLATE below. Do not copy it verbatim; adapt it.
- Exactly one specific observation drawn from the prospect's own website or notes.
- If the research text is empty, unusable, or clearly about a different company, DO NOT invent
  anything. Write the generic version of the template and say so in personalisation_note.
{rules}

Return ONLY valid JSON with these keys:
  "subject"              - plain text, under 70 characters, no clickbait, no emoji
  "body_html"            - the email body as simple HTML (<p> paragraphs only, no <html> wrapper,
                           no signature - a signature is appended separately)
  "personalisation_note" - ONE short sentence stating what specific detail you used, or
                           "GENERIC - no usable research" if you had nothing to work with

TEMPLATE:
---
{template}
---"""


def build_user_prompt(p: Prospect, research: str) -> str:
    lines = [
        "PROSPECT DETAILS",
        f"First name: {p.first_name or '(unknown)'}",
        f"Last name: {p.last_name or '(unknown)'}",
        f"Job title: {p.job_title or '(unknown)'}",
        f"Company: {p.company or '(unknown)'}",
    ]
    for label, value in p.context.items():
        lines.append(f"{label}: {value}")
    if p.notes:
        lines.append(f"Manual notes from the sender (treat as high-value, trust these): {p.notes}")
    if p.linkedin:
        lines.append(f"LinkedIn URL (reference only, not fetched): {p.linkedin}")
    lines.append("")
    lines.append("WEBSITE RESEARCH TEXT")
    lines.append(research.strip() if research.strip() else "(none available)")
    return "\n".join(lines)


# Closing words the templates' approved copy ends on. Checked against the whole trailing
# paragraph, not just its first word, so "Best Regards, Tooka" matches as one unit.
SIGNOFF_WORDS = frozenset({
    "best regards", "kind regards", "warm regards", "warmest regards", "regards",
    "many thanks", "best wishes", "best", "sincerely", "yours sincerely", "yours faithfully",
    "warmly", "thank you", "thanks",
})


def _is_signoff_line(text: str) -> bool:
    plain = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"[.,\s]+$", "", plain).strip().lower() in SIGNOFF_WORDS


def strip_signoff_html(body_html: str) -> str:
    """Drop a trailing sign-off paragraph the AI wrote into the body.

    Every template's approved copy ends with its own "Best regards, Tooka" as literal text,
    and the system prompt separately tells the AI not to include one because a signature is
    appended by the caller afterwards. In practice the AI keeps its own sign-off more often
    than not - about 5 times in 6 in a sample - because "follow the template's structure
    closely" wins out over "no signature". Left alone, every one of those emails would read
    "Best Regards, Tooka" twice. This is checked at the one point every caller passes
    through, so nothing downstream has to remember to do it.

    Handles the shape seen in practice - "<p>Best Regards,<br>Tooka</p>" as one paragraph -
    and the two-paragraph variant, "<p>Best regards,</p><p>Tooka</p>", in case the model
    formats it differently on another run.
    """
    text = (body_html or "").rstrip()
    paras = list(re.finditer(r"<p>(.*?)</p>", text, re.I | re.S))
    if not paras:
        return text

    last = paras[-1]
    lines = [l.strip() for l in re.split(r"<br\s*/?>", last.group(1), flags=re.I) if l.strip()]
    if lines and len(lines) <= 2 and _is_signoff_line(lines[0]):
        return text[: last.start()].rstrip()

    if len(paras) >= 2 and len(lines) == 1:
        name = re.sub(r"<[^>]+>", "", last.group(1)).strip()
        name_like = bool(name) and len(name.split()) <= 3 and not name.endswith((".", "?", "!"))
        prev_lines = [l.strip() for l in re.split(r"<br\s*/?>", paras[-2].group(1), flags=re.I)
                      if l.strip()]
        if name_like and len(prev_lines) == 1 and _is_signoff_line(prev_lines[0]):
            return text[: paras[-2].start()].rstrip()

    return text


def call_gemini(cfg: dict, system_prompt: str, user_prompt: str) -> dict[str, str]:
    ai = cfg["ai"]
    key = os.environ.get(ai["api_key_env"], "").strip()
    if not key:
        raise RuntimeError(
            f"No API key. Set the {ai['api_key_env']} environment variable "
            f"(get a free key at https://aistudio.google.com/apikey)."
        )
    base = ai.get("base_url") or GEMINI_BASE
    url = f"{base}/models/{ai['model']}:generateContent"
    payload = {
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"role": "user", "parts": [{"text": user_prompt}]}],
        "generationConfig": {
            "temperature": 0.7,
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "OBJECT",
                "properties": {
                    "subject": {"type": "STRING"},
                    "body_html": {"type": "STRING"},
                    "personalisation_note": {"type": "STRING"},
                },
                "required": ["subject", "body_html", "personalisation_note"],
            },
        },
    }
    r = requests.post(
        url,
        headers={"x-goog-api-key": key, "Content-Type": "application/json"},
        json=payload,
        timeout=90,
    )
    if r.status_code == 404:
        raise RuntimeError(
            f"Model '{ai['model']}' not found. Run 'python prospect_drafter.py models' "
            f"to list the models your key can use, then set ai.model in the config."
        )
    if r.status_code == 429:
        raise RuntimeError("Rate limited by the free tier. Increase ai.delay_seconds and retry.")
    r.raise_for_status()
    data = r.json()
    try:
        text = data["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError):
        raise RuntimeError(f"Unexpected AI response: {json.dumps(data)[:400]}")
    parsed = json.loads(text)
    return {
        "subject": str(parsed.get("subject", "")).strip(),
        "body_html": strip_signoff_html(str(parsed.get("body_html", "")).strip()),
        "personalisation_note": str(parsed.get("personalisation_note", "")).strip(),
    }


def list_models(cfg: dict) -> None:
    ai = cfg["ai"]
    key = os.environ.get(ai["api_key_env"], "").strip()
    if not key:
        sys.exit(f"Set the {ai['api_key_env']} environment variable first.")
    base = ai.get("base_url") or GEMINI_BASE
    r = requests.get(f"{base}/models", headers={"x-goog-api-key": key}, timeout=30)
    r.raise_for_status()
    print("Models your key can use for generateContent:\n")
    for m in r.json().get("models", []):
        if "generateContent" in m.get("supportedGenerationMethods", []):
            print(f"  {m['name'].removeprefix('models/'):45s} {m.get('displayName','')}")
    print("\nSet the one you want as ai.model in your config.")


def template_copy_block(template: str) -> str:
    """Just the email copy out of a template file, without the notes written for the AI.

    A template is two things in one file: the approved copy, and the guidance explaining how
    to fill it in. Only the copy belongs in an email. The copy is the section fenced by "---"
    lines; anything before the first fence and after the second is guidance. Without this,
    the no-AI fallback pastes the whole instruction file into the draft.
    """
    parts = re.split(r"(?m)^\s*-{3,}\s*$", template)
    block = parts[1] if len(parts) >= 2 else template
    # Drop the "APPROVED COPY - follow this closely:" style heading that opens the section.
    block = re.sub(
        r"(?is)^\s*(approved|suggested)\s+copy\b[^\n]*\n", "", block.strip(), count=1
    )
    return re.sub(r"\n{3,}", "\n\n", block).strip()


def strip_ai_instructions(text: str) -> str:
    """Replace the long bracketed passages written for the AI with a visible gap marker.

    Run this only after short placeholders like [COMPANY] have been substituted: an
    instruction block that still contains one would be nested brackets, which this
    deliberately simple pattern will not match.
    """
    return re.sub(r"\[[^\[\]]{40,}\]", "[FILL THIS IN]", text, flags=re.S)


def fallback_fill(template: str, p: Prospect, cfg: dict) -> dict[str, str]:
    """Template-only fill, used when the AI step fails."""
    s = cfg["sender"]
    mapping = {
        "first_name": p.first_name or "there",
        "last_name": p.last_name,
        "full_name": p.full_name,
        "company": p.company,
        "job_title": p.job_title,
        "your_name": s["your_name"],
        "your_company": s["your_company"],
    }
    body = template_copy_block(template)
    for k, v in mapping.items():
        body = body.replace("{{" + k + "}}", v)
    body = re.sub(r"\{\{[^}]+\}\}", "", body)
    # Templates write the company as [COMPANY] in prose the AI is meant to adapt. Substitute it
    # before stripping instructions, so a block containing it is no longer nested brackets.
    body = re.sub(r"\[COMPANY\]", p.company or "your firm", body, flags=re.I)
    body = strip_ai_instructions(body)
    # A stage's sign-off is appended separately, so drop it from the body.
    body = re.sub(
        r"(?is)\n\s*(best regards|kind regards|regards)\s*,?\s*\n.*$", "", body
    ).strip()
    # Templates are hard-wrapped for reading. Unwrap inside a paragraph, or the line breaks
    # land mid-sentence in the email.
    paragraphs = [
        f"<p>{escape_html(re.sub(r'[ \t]*\n[ \t]*', ' ', x.strip()))}</p>"
        for x in body.split("\n\n") if x.strip()
    ]
    return {
        "subject": f"{s['your_company']} <> {p.company}".strip(" <>"),
        "body_html": "\n".join(paragraphs),
        "personalisation_note": "GENERIC - template fallback, AI step unavailable",
    }


def escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
    )


# ---------------------------------------------------------------------------
# Review file
# ---------------------------------------------------------------------------

def write_review_file(results: list[DraftResult], cfg: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime("%d %b %Y, %H:%M")
    generic = sum(
        1
        for r in results
        if not r.research_ok or r.personalisation_note.upper().startswith("GENERIC")
    )

    head = [
        "# Draft review",
        "",
        f"Generated {now} — {len(results)} drafts, {generic} with no usable research.",
        "",
        "## How to use this file",
        "",
        "1. Read each draft below. Edit the SUBJECT and BODY freely.",
        "2. Set STATUS to `APPROVED` for the ones you want, or `SKIP` to leave them out.",
        "3. Save the file, then run:  `python prospect_drafter.py push`",
        "",
        "Nothing is sent. Approved drafts are created as unsent drafts only.",
        "BODY is HTML — keep the `<p>` tags if you want paragraph spacing.",
        "",
        REVIEW_SEPARATOR,
        "",
    ]
    blocks: list[str] = []
    for i, r in enumerate(results, 1):
        p = r.prospect
        flags = []
        if not r.research_ok:
            flags.append("NO RESEARCH")
        if not r.ai_ok:
            flags.append("AI FALLBACK")
        flag_str = ("  ⚠ " + ", ".join(flags)) if flags else ""
        block = [
            f"## [{i}] {p.full_name or '(no name)'} — {p.company or '(no company)'}{flag_str}",
            "",
            f"- TO: {p.email}",
            f"- ROW: {p.row}",
            f"- WEBSITE: {p.website or '(none)'}",
            f"- WHY THIS ANGLE: {r.personalisation_note or '(none)'}",
        ]
        if r.warnings:
            block.append(f"- NOTES: {'; '.join(r.warnings)}")
        block += [
            "",
            "STATUS: APPROVED",
            "",
            f"SUBJECT: {r.subject}",
            "",
            "BODY:",
            "```html",
            r.body_html.strip(),
            "```",
            "",
            REVIEW_SEPARATOR,
            "",
        ]
        blocks.append("\n".join(block))

    path.write_text("\n".join(head) + "\n".join(blocks), encoding="utf-8")


REVIEW_ENTRY_RE = re.compile(
    r"^##\s*\[(?P<idx>\d+)\][^\n]*\n(?P<body>.*?)(?=^##\s*\[|\Z)",
    re.S | re.M,
)


def parse_review_file(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        sys.exit(f"No review file at {path}. Run the research phase first.")
    text = path.read_text(encoding="utf-8")
    entries: list[dict[str, str]] = []
    for m in REVIEW_ENTRY_RE.finditer(text):
        chunk = m.group("body")

        def grab(label: str) -> str:
            mm = re.search(rf"^[-\s]*{label}:\s*(.*)$", chunk, re.M)
            return mm.group(1).strip() if mm else ""

        body_m = re.search(r"^BODY:\s*\n```(?:html)?\s*\n(.*?)\n```", chunk, re.S | re.M)
        entries.append(
            {
                "index": m.group("idx"),
                "to": grab("TO"),
                "row": grab("ROW"),
                "status": grab("STATUS").upper(),
                "subject": grab("SUBJECT"),
                "body_html": body_m.group(1).strip() if body_m else "",
                "note": grab("WHY THIS ANGLE"),
            }
        )
    return entries


# ---------------------------------------------------------------------------
# Outlook backends
# ---------------------------------------------------------------------------

def push_eml(entry: dict, cfg: dict) -> str:
    """Write a .eml file that Outlook opens as an unsent, editable message."""
    from email.message import EmailMessage

    out_dir = Path(cfg["output"]["eml_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)

    msg = EmailMessage()
    msg["Subject"] = entry["subject"]
    msg["To"] = entry["to"]
    # X-Unsent tells Outlook to open this as a composable draft with a Send button.
    msg["X-Unsent"] = "1"
    html = entry["body_html"] + "\n" + cfg["sender"].get("signature_html", "")
    msg.set_content(strip_tags(html))
    msg.add_alternative(f"<html><body>{html}</body></html>", subtype="html")

    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", entry["to"] or f"row{entry['row']}")[:60]
    fp = out_dir / f"{int(entry['index']):03d}_{safe}.eml"
    fp.write_bytes(msg.as_bytes())
    return str(fp)


def strip_tags(html: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.I)
    text = re.sub(r"</p>", "\n\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def push_com(entry: dict, cfg: dict) -> str:
    """Classic Outlook desktop on Windows. No authentication needed."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError(
            "The 'com' backend needs classic Outlook for Windows and pywin32.\n"
            "Install with:  pip install pywin32\n"
            "Note: the new Outlook for Windows does NOT support this. Use 'eml' or 'graph'."
        )
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # olMailItem
    mail.To = entry["to"]
    mail.Subject = entry["subject"]
    mail.HTMLBody = entry["body_html"] + "\n" + cfg["sender"].get("signature_html", "")
    mail.Save()  # lands in Drafts, unsent
    return "Outlook Drafts"


# --- Microsoft Graph (device code flow) ------------------------------------

def graph_token(cfg: dict, interactive: bool = True) -> str:
    oc = cfg["outlook"]
    client_id = oc.get("graph_client_id", "").strip()
    if not client_id:
        raise RuntimeError(
            "No graph_client_id in the config.\n"
            "The Graph backend needs an Entra app registration. See IT-REQUEST.md — it is a\n"
            "two-minute job for whoever administers your tenant. Until then use --backend eml."
        )
    tenant = oc.get("graph_tenant_id") or "organizations"
    cache_path = Path(oc["token_cache"])

    cached = {}
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            cached = {}

    if cached.get("access_token") and cached.get("expires_at", 0) > time.time() + 120:
        return cached["access_token"]

    if cached.get("refresh_token"):
        tok = _graph_refresh(tenant, client_id, cached["refresh_token"])
        if tok:
            _save_token(cache_path, tok)
            return tok["access_token"]

    if not interactive:
        raise RuntimeError("Not signed in. Run:  python prospect_drafter.py auth")

    tok = _graph_device_code(tenant, client_id)
    _save_token(cache_path, tok)
    return tok["access_token"]


def _graph_device_code(tenant: str, client_id: str) -> dict:
    r = requests.post(
        f"{LOGIN_BASE}/{tenant}/oauth2/v2.0/devicecode",
        data={"client_id": client_id, "scope": GRAPH_SCOPES},
        timeout=30,
    )
    r.raise_for_status()
    dc = r.json()
    print("\n" + "-" * 70)
    print(dc.get("message", f"Go to {dc['verification_uri']} and enter {dc['user_code']}"))
    print("-" * 70 + "\n")

    interval = int(dc.get("interval", 5))
    deadline = time.time() + int(dc.get("expires_in", 900))
    while time.time() < deadline:
        time.sleep(interval)
        tr = requests.post(
            f"{LOGIN_BASE}/{tenant}/oauth2/v2.0/token",
            data={
                "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
                "client_id": client_id,
                "device_code": dc["device_code"],
            },
            timeout=30,
        )
        body = tr.json()
        if tr.status_code == 200:
            print("Signed in.\n")
            return body
        err = body.get("error")
        if err == "authorization_pending":
            continue
        if err == "slow_down":
            interval += 5
            continue
        raise RuntimeError(f"Sign-in failed: {body.get('error_description', err)}")
    raise RuntimeError("Sign-in timed out.")


def _graph_refresh(tenant: str, client_id: str, refresh_token: str) -> dict | None:
    r = requests.post(
        f"{LOGIN_BASE}/{tenant}/oauth2/v2.0/token",
        data={
            "grant_type": "refresh_token",
            "client_id": client_id,
            "refresh_token": refresh_token,
            "scope": GRAPH_SCOPES,
        },
        timeout=30,
    )
    return r.json() if r.status_code == 200 else None


def _save_token(path: Path, tok: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tok = dict(tok)
    tok["expires_at"] = time.time() + int(tok.get("expires_in", 3600))
    path.write_text(json.dumps(tok), encoding="utf-8")
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def push_graph(entry: dict, cfg: dict, token: str) -> str:
    payload = {
        "subject": entry["subject"],
        "body": {
            "contentType": "HTML",
            "content": entry["body_html"] + "\n" + cfg["sender"].get("signature_html", ""),
        },
        "toRecipients": [{"emailAddress": {"address": entry["to"]}}],
    }
    r = requests.post(
        f"{GRAPH_BASE}/me/messages",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=payload,
        timeout=45,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Graph error {r.status_code}: {r.text[:300]}")
    return r.json().get("id", "created")


# ---------------------------------------------------------------------------
# State
# ---------------------------------------------------------------------------

DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d",
    "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y", "%d/%m/%y",
)


def parse_date(value) -> date | None:
    """Parse a spreadsheet date cell. Day-first for ambiguous slash dates (UK convention)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def find_sheet(names: list[str], want: str | None) -> str | None:
    """The real sheet name matching `want`, or None.

    Case- and whitespace-insensitive: a sheet renamed by hand often picks up a trailing
    space, and "First Contact Date " should still match "First Contact Date".
    """
    if not want:
        return None
    key = re.sub(r"\s+", " ", str(want)).strip().lower()
    for n in names:
        if re.sub(r"\s+", " ", str(n)).strip().lower() == key:
            return n
    return None


def resolve_sheet_name(names: list[str], want: str | None, active: str | None = None) -> str:
    """Which sheet to actually read, given the one the workflow asks for.

    Order of preference: the sheet the workflow is configured for, then whichever sheet
    Excel had active when the file was saved, then the first one.
    """
    if not names:
        raise ValueError("This workbook has no sheets.")
    return find_sheet(names, want) or find_sheet(names, active) or names[0]


def infer_touches(status: str, explicit: str, patterns: dict | None) -> int:
    """How many emails this person has already had.

    `explicit` is the Touches column, which is authoritative when present. Some sheets do
    not have one and record contact as free text in Status instead - "Email sent - Linkedin
    Sent". `patterns` maps a touch count to the phrases that imply it, matched anywhere in
    the status rather than only at the start, and the highest matching count wins. Without
    this, a sheet full of "Email sent" rows reads as nobody having been contacted, and the
    whole list gets queued for a first email it has already had.
    """
    raw = (explicit or "").strip()
    if raw:
        try:
            return int(float(raw))
        except ValueError:
            pass

    s = re.sub(r"\s+", " ", (status or "").strip().lower())
    if not s:
        return 0

    best = 0
    for count, phrases in (patterns or {}).items():
        try:
            n = int(count)
        except (TypeError, ValueError):
            continue
        if n <= best:
            continue
        if any(re.sub(r"\s+", " ", str(p).strip().lower()) in s for p in phrases):
            best = n
    if best:
        return best

    # Long-standing fallback: sheets written by an earlier version of this tool recorded the
    # first touch as "Drafted 17 Aug 2026" in Status.
    return 1 if re.match(r"^(drafted|sent|contacted)", s) else 0


def stop_reason(status: str, sc: dict) -> str:
    """Why this row should never be contacted, or "" if it is fine.

    Two mechanisms, because sheets record this two ways. `stop_statuses` matches the start
    of a controlled value ("Replied - Positive"). `stop_contains` matches anywhere, for
    free-text judgement notes like "Email sent - Family Portraits - too small".
    """
    s = re.sub(r"\s+", " ", (status or "").strip().lower())
    if not s:
        return ""
    if should_skip_status(status, sc.get("stop_statuses", [])):
        return f"status is '{status.strip()}'"
    for phrase in sc.get("stop_contains", []) or []:
        p = re.sub(r"\s+", " ", str(phrase).strip().lower())
        if p and p in s:
            return f"status says '{phrase}'"
    return ""


def normalise_score(value) -> str:
    """Fit scores arrive from Excel as 4, 4.0 or "4". Reduce them all to "4"."""
    s = str(value if value is not None else "").strip()
    if not s:
        return ""
    try:
        f = float(s)
    except ValueError:
        return s
    return str(int(f)) if f.is_integer() else str(f)


def normalise_status(value) -> str:
    """Case- and whitespace-insensitive form of a status, for comparing against config."""
    return re.sub(r"\s+", " ", str(value if value is not None else "").strip()).lower()


def blocked_by_column(p: "Prospect", sc: dict) -> str:
    """Reason this row is off limits because of another column, or "" if it is fine.

    Configured as spreadsheet.skip_when, e.g. {"Interested?": ["NO"]}. The cold database
    uses this: a row can be marked "Interested? = NO" while its Status still reads
    "First Contact", and without this check the follow-ups would keep going out to
    someone who has already said no.
    """
    for header, values in (sc.get("skip_when") or {}).items():
        actual = normalise_status(p.context.get(str(header).strip(), ""))
        if actual and actual in {normalise_status(v) for v in values}:
            return f"{header} is '{p.context.get(str(header).strip())}'"
    return ""


def _wait_gate(p: "Prospect", cfg: dict, today: date, require_date: bool) -> tuple[bool, str]:
    wait = int(cfg.get("sequence", {}).get("wait_days", 7))
    last = p.latest_contact
    if last is None:
        # Status-gated rows are trusted: the Status column already says the previous email
        # went out, so a missing date is a gap in the sheet rather than a reason to skip.
        return (False, "no contact date recorded") if require_date else (True, "")
    waited = (today - last).days
    if waited < wait:
        return False, f"only {waited} day(s) since last contact (needs {wait})"
    return True, ""


def eligible_by_status(
    p: "Prospect", stage: int, cfg: dict, today: date | None = None
) -> tuple[bool, str]:
    """Status-column version of the stage gate, used by the Cold Call workflow.

    The Status column is the single source of truth for where a prospect is in the
    sequence: blank means nobody has written yet, "First Contact" means email 1 has gone,
    and so on. `sequence.status_flow` holds that mapping.
    """
    today = today or date.today()
    sc = cfg["spreadsheet"]
    seq = cfg.get("sequence", {})

    if not p.email:
        return False, "no email address"
    if reason := stop_reason(p.status, sc):
        return False, f"{reason} — this row is closed"
    if blocker := blocked_by_column(p, sc):
        return False, blocker

    flow = (seq.get("status_flow") or {}).get(str(stage))
    if not flow:
        return False, f"no status rule is configured for email {stage}"

    allowed = {normalise_status(x) for x in flow.get("from", [])}
    if normalise_status(p.status) not in allowed:
        shown = ", ".join(f"'{x}'" if x else "blank" for x in flow.get("from", []))
        current = f"'{p.status}'" if p.status else "blank"
        return False, f"status is {current} — email {stage} needs {shown}"

    if stage > 1:
        return _wait_gate(p, cfg, today, require_date=False)
    return True, ""


def eligible_for_stage(
    p: "Prospect", stage: int, cfg: dict, today: date | None = None
) -> tuple[bool, str]:
    """Can this prospect receive email number `stage`? Returns (yes/no, reason if no)."""
    today = today or date.today()
    sc = cfg["spreadsheet"]

    if str(cfg.get("sequence", {}).get("gate", "touches")).lower() == "status":
        return eligible_by_status(p, stage, cfg, today)

    if not p.email:
        return False, "no email address"
    if reason := stop_reason(p.status, sc):
        return False, reason
    if blocker := blocked_by_column(p, sc):
        return False, blocker
    if p.touches != stage - 1:
        if p.touches >= stage:
            return False, f"already had {p.touches} email(s)"
        return False, f"only {p.touches} email(s) so far — not ready for number {stage}"
    if stage > 1:
        return _wait_gate(p, cfg, today, require_date=True)
    return True, ""


# ---------------------------------------------------------------------------
# Workflows
# ---------------------------------------------------------------------------

def workflow_names(cfg: dict) -> list[str]:
    """Configured workflow keys, with the default one first."""
    names = list((cfg.get("workflows") or {}).keys())
    default = cfg.get("default_workflow")
    if default in names:
        names.remove(default)
        names.insert(0, default)
    return names


def workflow_label(cfg: dict, name: str) -> str:
    return ((cfg.get("workflows") or {}).get(name) or {}).get("label", name)


def apply_workflow(cfg: dict, name: str) -> dict:
    """Return `cfg` with the named workflow's settings merged over the top.

    Each workflow block is an overlay on the shared config, so anything it does not
    mention - the sender details, the AI settings, the Outlook setup - stays as it is.
    The workflow's own metadata is kept under cfg["workflow"] for the app to read.
    """
    block = dict(((cfg.get("workflows") or {}).get(name) or {}))
    if not block:
        return cfg
    overlay = {k: v for k, v in block.items()
               if k not in ("label", "blurb", "fit_score", "writeback",
                            "called_first_default")}
    merged = deep_merge(cfg, overlay)
    merged["workflow"] = {
        "name": name,
        "label": block.get("label", name),
        "blurb": block.get("blurb", ""),
        "fit_score": block.get("fit_score") or {},
        "writeback": block.get("writeback") or {},
        "called_first_default": bool(block.get("called_first_default", False)),
    }
    return merged


def next_status(cfg: dict, stage: int) -> str:
    """The Status value to write once email `stage` has been drafted."""
    flow = (cfg.get("sequence", {}).get("status_flow") or {}).get(str(stage)) or {}
    return str(flow.get("to") or "").strip()


def fit_score_values(prospects: list["Prospect"], cfg: dict) -> list[str]:
    """Scores actually present in the sheet, ordered the way the config lists them."""
    present = {p.fit_score for p in prospects if p.fit_score}
    ordered = [s for s in (cfg.get("workflow", {}).get("fit_score", {}).get("options") or [])
               if s in present]
    return ordered + sorted(present - set(ordered))


def stage_signature(cfg: dict, stage: int) -> str:
    """The sign-off for this stage.

    Approved copy does not use the same sign-off throughout a sequence: the cold first email
    ends "Best Regards" and the follow-ups end "Kind regards". Where a workflow sets
    sequence.signatures, the stage wins; otherwise this is the sender's single signature and
    nothing changes.
    """
    sigs = cfg.get("sequence", {}).get("signatures") or {}
    return str(sigs.get(str(stage)) or cfg.get("sender", {}).get("signature_html", ""))


def stage_template_path(cfg: dict, stage: int, base_dir: Path | None = None) -> Path:
    templates = cfg.get("sequence", {}).get("templates", {})
    name = templates.get(str(stage)) or cfg.get("template_path", "template.md")
    p = Path(name)
    if base_dir and not p.is_absolute():
        p = base_dir / p
    return p


def should_skip_status(status: str, skip_values) -> bool:
    """True if the row's Status marks it as already handled.

    Matches on prefix, so a status written back as "Drafted 17 Aug 2026" still
    matches the configured "drafted".
    """
    s = (status or "").strip().lower()
    if not s:
        return False
    return any(s.startswith(str(v).strip().lower()) for v in skip_values if str(v).strip())


def load_state(cfg: dict) -> dict:
    p = Path(cfg["output"]["state_file"])
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}


def save_state(cfg: dict, state: dict) -> None:
    p = Path(cfg["output"]["state_file"])
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(state, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_research(cfg: dict, args) -> None:
    template_path = Path(cfg["template_path"])
    if not template_path.exists():
        sys.exit(f"No email template at {template_path}.")
    template = template_path.read_text(encoding="utf-8")

    prospects = read_prospects(cfg)
    state = load_state(cfg)
    skip_statuses = cfg["spreadsheet"]["skip_when_status_in"]

    queue: list[Prospect] = []
    for p in prospects:
        if should_skip_status(p.status, skip_statuses):
            continue
        if state.get(p.key, {}).get("pushed") and not args.redo:
            continue
        if not p.email:
            print(f"  row {p.row}: no email address, skipped")
            continue
        queue.append(p)

    if args.limit:
        queue = queue[: args.limit]

    if not queue:
        print("Nothing to do. Every prospect is either already drafted or filtered out.")
        return

    print(f"Researching and drafting {len(queue)} prospect(s).\n")
    system_prompt = build_system_prompt(cfg, template)
    results: list[DraftResult] = []

    for i, p in enumerate(queue, 1):
        label = f"[{i}/{len(queue)}] {p.full_name or p.email} ({p.company})"
        print(f"{label} ... ", end="", flush=True)
        res = DraftResult(prospect=p)

        research_text = ""
        if cfg["research"]["enabled"] and p.website:
            research_text, warns = fetch_site_text(p.website, cfg)
            res.warnings.extend(warns)
            res.research_ok = len(research_text) >= 200
        elif not p.website:
            res.warnings.append("No website in the spreadsheet.")

        try:
            out = call_gemini(cfg, system_prompt, build_user_prompt(p, research_text))
            res.ai_ok = True
        except Exception as exc:  # noqa: BLE001
            res.warnings.append(f"AI step failed: {exc}")
            out = fallback_fill(template, p, cfg)

        res.subject = out["subject"]
        res.body_html = out["body_html"]
        res.personalisation_note = out["personalisation_note"]
        results.append(res)

        marks = "ok" if (res.research_ok and res.ai_ok) else "ok (with warnings)"
        print(marks)

        if i < len(queue):
            time.sleep(float(cfg["ai"]["delay_seconds"]))

    review_path = Path(cfg["output"]["review_file"])
    write_review_file(results, cfg, review_path)

    weak = sum(
        1
        for r in results
        if not r.research_ok or not r.ai_ok
        or r.personalisation_note.upper().startswith("GENERIC")
    )
    print(f"\nWrote {len(results)} drafts to {review_path}")
    if weak:
        print(f"  {weak} are flagged (no research or AI fallback) — check those first.")
    print("\nNext: open that file, edit the copy, set STATUS to APPROVED or SKIP,")
    print("then run:  python prospect_drafter.py push")


def cmd_push(cfg: dict, args) -> None:
    backend = args.backend or cfg["outlook"]["backend"]
    entries = parse_review_file(Path(cfg["output"]["review_file"]))
    approved = [e for e in entries if e["status"] == "APPROVED"]
    skipped = len(entries) - len(approved)

    if not approved:
        sys.exit(f"No entries marked APPROVED in the review file ({len(entries)} found).")

    print(f"Pushing {len(approved)} approved draft(s) via '{backend}'. {skipped} skipped.\n")
    if args.dry_run:
        for e in approved:
            print(f"  would draft to {e['to']}: {e['subject']}")
        return

    token = graph_token(cfg) if backend == "graph" else None
    state = load_state(cfg)
    ok = 0

    for e in approved:
        if not e["to"]:
            print(f"  [{e['index']}] no recipient, skipped")
            continue
        try:
            if backend == "eml":
                where = push_eml(e, cfg)
            elif backend == "com":
                where = push_com(e, cfg)
            elif backend == "graph":
                where = push_graph(e, cfg, token)
            else:
                sys.exit(f"Unknown backend '{backend}'. Use eml, com or graph.")
            key = e["to"].strip().lower()
            state[key] = {
                "pushed": True,
                "backend": backend,
                "subject": e["subject"],
                "at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "location": where,
            }
            ok += 1
            print(f"  [{e['index']}] {e['to']} → {where}")
        except Exception as exc:  # noqa: BLE001
            print(f"  [{e['index']}] FAILED for {e['to']}: {exc}")

    save_state(cfg, state)
    print(f"\n{ok}/{len(approved)} drafted.")
    if backend == "eml":
        print(f"Open the .eml files in {cfg['output']['eml_dir']} — double-click, or select all")
        print("and drag them into your Outlook Drafts folder. Each opens ready to send.")


def cmd_auth(cfg: dict, args) -> None:
    token = graph_token(cfg, interactive=True)
    r = requests.get(f"{GRAPH_BASE}/me", headers={"Authorization": f"Bearer {token}"}, timeout=30)
    if r.status_code == 200:
        me = r.json()
        print(f"Signed in as {me.get('displayName')} <{me.get('mail') or me.get('userPrincipalName')}>")
    else:
        print(f"Token acquired but /me returned {r.status_code}: {r.text[:200]}")


def cmd_status(cfg: dict, args) -> None:
    state = load_state(cfg)
    review = Path(cfg["output"]["review_file"])
    print(f"Config spreadsheet : {cfg['spreadsheet']['path']}")
    print(f"Review file        : {review} ({'present' if review.exists() else 'not generated yet'})")
    print(f"Already drafted    : {sum(1 for v in state.values() if v.get('pushed'))}")
    if review.exists():
        entries = parse_review_file(review)
        appr = sum(1 for e in entries if e["status"] == "APPROVED")
        print(f"In review file     : {len(entries)} total, {appr} marked APPROVED")


def cmd_init(cfg_path: Path) -> None:
    if cfg_path.exists():
        sys.exit(f"{cfg_path} already exists. Delete it first if you want a fresh one.")
    cfg_path.write_text(json.dumps(DEFAULT_CONFIG, indent=2), encoding="utf-8")
    tpl = Path(DEFAULT_CONFIG["template_path"])
    if not tpl.exists():
        tpl.write_text(SAMPLE_TEMPLATE, encoding="utf-8")
    print(f"Created {cfg_path} and {tpl}.")
    print("Edit both, set your GEMINI_API_KEY, then run:  python prospect_drafter.py research --limit 2")


SAMPLE_TEMPLATE = """Hi {{first_name}},

[Opening line: one specific, accurate observation about {{company}} drawn from their
website. Not flattery. Something that shows this was written for them.]

[Bridge: connect that observation to a problem Gateway Global solves. One or two
sentences. Concrete, no jargon, no buzzwords.]

[Proof: one short, credible line. A comparable client, a number, or a specific
outcome. Keep it modest and true.]

[Ask: a single low-friction question. A short call, or simply whether it is worth
a conversation.]
"""


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Research prospects and draft personalised emails into Outlook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--config", default="config.json", help="path to config file")
    sub = ap.add_subparsers(dest="command", required=True)

    sub.add_parser("init", help="write a starter config.json and template.md")
    sub.add_parser("models", help="list AI models your API key can use")
    sub.add_parser("auth", help="sign in to Microsoft Graph (device code)")
    sub.add_parser("status", help="show what has been drafted so far")

    r = sub.add_parser("research", help="phase 1: research and draft to the review file")
    r.add_argument("--limit", type=int, default=0, help="only process the first N prospects")
    r.add_argument("--redo", action="store_true", help="include prospects already pushed")

    p = sub.add_parser("push", help="phase 2: create Outlook drafts from the review file")
    p.add_argument("--backend", choices=["eml", "com", "graph"], help="override config backend")
    p.add_argument("--dry-run", action="store_true", help="show what would happen")

    args = ap.parse_args()
    cfg_path = Path(args.config)

    if args.command == "init":
        cmd_init(cfg_path)
        return

    cfg = load_config(cfg_path)
    {
        "models": lambda: list_models(cfg),
        "auth": lambda: cmd_auth(cfg, args),
        "status": lambda: cmd_status(cfg, args),
        "research": lambda: cmd_research(cfg, args),
        "push": lambda: cmd_push(cfg, args),
    }[args.command]()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(130)
    except RuntimeError as exc:
        print(f"\n{exc}", file=sys.stderr)
        sys.exit(1)
    except requests.RequestException as exc:
        print(f"\nNetwork problem: {exc}", file=sys.stderr)
        sys.exit(1)
