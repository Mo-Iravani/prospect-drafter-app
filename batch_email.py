"""
Batch Email — one static email, merged with a list.

The other two workflows write a different email for every prospect. This one does the
opposite: the wording is fixed and already approved, and only the changing bits — a name,
a company, a reference number — come from the spreadsheet. There is no website research
and no AI anywhere in this path, so a hundred drafts render instantly and read exactly as
they were written.

The whole design rests on one rule: **a row with a hole in it never becomes a draft.**
At a hundred recipients the failure that matters is not a clumsy sentence, it is a real
email to a real person opening "Dear ," and quoting reference "". So every placeholder has
to resolve to a mapped column with something in it, and any row that cannot manage that is
blocked and named rather than quietly papered over.

Placeholders are written {{Column Name}} — the spreadsheet's own heading, spelled however
suits. Matching is case- and whitespace-insensitive and ignores a trailing colon, because
"{{Reference Number}}", "{{reference number}}" and "{{Reference Number:}}" are plainly the
same request. A trailing question mark, {{Job Title?}}, marks one optional: blank is then
allowed and the line closes up rather than the row being blocked.

The subject line is templated the same way as the body, from the same columns: it is just
another string put through the same substitution.
"""

from __future__ import annotations

import csv
import html
import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import prospect_drafter as pd_lib

PLACEHOLDER_RE = re.compile(r"\{\{\s*([^{}]*?)\s*\}\}")

# Deliberately loose. This is a typo catcher, not an RFC 5322 parser: the job is to stop
# "j.smith@acme" and "see notes" reaching Outlook, not to adjudicate exotic-but-legal
# addresses. One @, no whitespace or list separators, a dotted domain — anything else is
# left to the mail server to argue about.
EMAIL_RE = re.compile(r"^[^@\s,;<>]+@[^@\s,;<>]+\.[A-Za-z]{2,}$")

MAX_BLANK_RUN = 200  # stop reading after this many consecutive empty rows


# ---------------------------------------------------------------------------
# Names
# ---------------------------------------------------------------------------

def norm(name: object) -> str:
    """The comparison form of a column or placeholder name.

    Collapsing whitespace and dropping a trailing colon matters more than it looks: a
    heading typed by hand picks up a double space or a trailing one constantly, and a
    placeholder copied out of a sentence often keeps the colon that followed it. The
    trailing "?" that marks a placeholder optional goes the same way, so {{Job Title?}} and
    a column called "Job Title" are recognised as the same thing.
    """
    return re.sub(r"\s+", " ", str(name or "").strip()).strip(":?").strip().lower()


def _squash(name: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(name))


def find_placeholders(*texts: str) -> list[str]:
    """Placeholder names in order of first appearance, de-duplicated case-insensitively.

    The original spelling of the first occurrence is what comes back, because that is what
    the user typed and what they will be looking for on screen.
    """
    seen: set[str] = set()
    out: list[str] = []
    for text in texts:
        for m in PLACEHOLDER_RE.finditer(text or ""):
            raw = m.group(1).strip()
            key = norm(raw)
            if not key or key in seen:
                continue
            seen.add(key)
            out.append(raw)
    return out


def has_placeholder(text: str) -> bool:
    """True if any {{...}} survived substitution. The last gate before Outlook."""
    return bool(PLACEHOLDER_RE.search(text or ""))


def looks_like_signoff(text: str) -> bool:
    """True if the template already ends on its own "Kind regards, Tooka".

    A pasted email nearly always carries its sign-off with it, and the app appends a
    signature of its own. Left alone that reads "Kind regards, Tooka" twice, which is the
    single most obvious tell that an email was machine-made.
    """
    paras = [p for p in re.split(r"\n\s*\n", (text or "").strip()) if p.strip()]
    if not paras:
        return False
    lines = [l.strip() for l in paras[-1].splitlines() if l.strip()]
    if not lines or len(lines) > 3:
        return False
    first = re.sub(r"[.,\s]+$", "", lines[0]).strip().lower()
    return first in pd_lib.SIGNOFF_WORDS


# ---------------------------------------------------------------------------
# Links
# ---------------------------------------------------------------------------
#
# Every draft in this app travels as plain text, from the box the user reads it in right up to
# the moment it becomes HTML for Outlook. Plain text has nowhere to keep a link's target: a
# textarea cannot hold formatting, so a hyperlink pasted out of Word or Outlook arrives as its
# visible words with the address silently gone.
#
# So links are written down instead, in the one notation people already recognise from chat
# apps and README files:
#
#     [our membership page](https://example.com/join)   text and target, kept together
#     https://example.com/join                          a bare address, linked as it stands
#
# Both become real <a> tags on the way into Outlook, and a link that comes back out of HTML is
# written back in the same notation, so nothing is lost by editing a draft and nothing is lost
# by uploading a Word file that had a proper hyperlink in it.

# A {{placeholder}} is allowed inside an address, so a per-person link like
# https://portal.example.com/renew?ref={{Ref No.}} holds together in the box the user reads
# it in — the space inside the braces would otherwise look like the end of the address. By
# the time a draft is turned into HTML the placeholder has already been substituted, so this
# only has to survive being looked at.
_URL = r"(?:https?://|mailto:|www\.)(?:\{\{[^}\n]*\}\}|[^\s<>\"'\]])+"

# One regex, one left-to-right pass: the [text](url) form is listed first so it wins wherever
# it applies, and the bare-address form cannot then chew up the target inside it.
#
# The target of a [text](target) link may also be a placeholder on its own — [{{Company}}]
# ({{Website}}) — so a whole address can come out of a column. A bare {{Website}} sitting in a
# sentence is deliberately not treated as a link: it is just a value being printed.
LINK_RE = re.compile(
    rf"\[([^\]\n]+)\]\(\s*({_URL}|\{{\{{[^}}\n]*\}}\}})\s*\)|({_URL})", re.I
)

# Punctuation that ends a sentence rather than the address, so "see https://x.com." links to
# the right place. A closing bracket only counts as trailing if it is unmatched.
_TRAILING = ".,;:!?"


def _split_trailing(url: str) -> tuple[str, str]:
    tail = ""
    while url and (url[-1] in _TRAILING or (url[-1] == ")" and url.count("(") < url.count(")"))):
        tail = url[-1] + tail
        url = url[:-1]
    return url, tail


def _href(url: str) -> str:
    """A bare www.example.com needs a scheme or the browser reads it as a relative path."""
    return f"http://{url}" if url.lower().startswith("www.") else url


def find_links(text: str) -> list[tuple[str, str]]:
    """The (visible text, target) pairs in a template, for showing the user what was found."""
    out: list[tuple[str, str]] = []
    for m in LINK_RE.finditer(text or ""):
        if m.group(2):
            url, _ = _split_trailing(m.group(2))
            out.append((m.group(1).strip(), url))
        else:
            url, _ = _split_trailing(m.group(3))
            out.append((url, url))
    return out


def linkify(escaped_html: str) -> str:
    """Turn the written-down links in already-escaped text into real anchors.

    Takes text that has been through HTML escaping, so the only markup present is the <br>
    that escaping produced. That ordering is deliberate: escaping first means a stray < or &
    in the wording is safe, and matching afterwards means the anchors it inserts are not
    escaped in turn.
    """
    def replace(m: re.Match) -> str:
        if m.group(2):
            url, tail = _split_trailing(m.group(2))
            label = m.group(1).strip()
        else:
            url, tail = _split_trailing(m.group(3))
            label = url
        if not url:
            return m.group(0)
        return f'<a href="{_href(url)}">{label}</a>{tail}'

    return LINK_RE.sub(replace, escaped_html or "")


def plain_links(text: str) -> str:
    """Written-down links flattened for somewhere that cannot hold a real one.

    The plain-text half of an email, and the Outlook compose link, are both plain text.
    "[our page](https://x)" would show up there verbatim, so it becomes "our page
    (https://x)" instead — which every mail client turns back into a working link anyway.
    """
    def replace(m: re.Match) -> str:
        if not m.group(2):
            return m.group(0)
        url, tail = _split_trailing(m.group(2))
        label = m.group(1).strip()
        return (url if label == url else f"{label} ({url})") + tail

    return LINK_RE.sub(replace, text or "")


def delinkify(html: str) -> str:
    """The reverse: an <a> tag written back as [text](url), so editing a draft keeps its links."""
    def replace(m: re.Match) -> str:
        url = m.group(1).strip()
        label = re.sub(r"<[^>]+>", "", m.group(2)).strip()
        if not label or label == url or _href(label) == url:
            return url
        return f"[{label}]({url})"

    return re.sub(
        r"<a\b[^>]*?href=[\"']([^\"']*)[\"'][^>]*>(.*?)</a\s*>", replace, html or "",
        flags=re.I | re.S,
    )


# ---------------------------------------------------------------------------
# The list
# ---------------------------------------------------------------------------

def cell_to_text(value: object) -> str:
    """A spreadsheet cell as it should read inside a sentence.

    str() is not good enough here. A reference number stored as a number arrives as 40521.0
    and would go out in the email with the .0 still on it; a date arrives as a datetime and
    would go out as "2026-08-21 00:00:00". Internal line breaks are kept — a cell holding
    an address is meant to have them — but runs of spaces and tabs are collapsed, which
    quietly fixes the trailing-space-in-the-cell problem no amount of validation catches.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        if value.hour or value.minute:
            return value.strftime("%d %B %Y, %H:%M")
        return value.strftime("%d %B %Y")
    if isinstance(value, date):
        return value.strftime("%d %B %Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"[ \t]+", " ", str(value)).strip()


@dataclass
class BatchRow:
    row: int                        # worksheet row number, so a problem can be named
    values: dict[str, str]          # normalised column name -> cell text


@dataclass
class Table:
    headers: list[str]              # original spelling, sheet order, duplicates dropped
    rows: list[BatchRow]
    duplicate_headers: list[str] = field(default_factory=list)

    @property
    def keys(self) -> list[str]:
        return [norm(h) for h in self.headers]

    def header_for(self, key: str) -> str:
        """The user's spelling of a normalised column name, for showing on screen."""
        for h in self.headers:
            if norm(h) == key:
                return h
        return key


def read_table(path: Path, sheet_name: str | None = None, header_row: int = 1) -> Table:
    """Read a sheet as headings plus rows of text, with no expected schema at all.

    This is the one real difference from the other two workflows. They map fixed fields
    (company, first_name, ...) named in config.json. A batch list's columns are whatever
    this batch happens to have, so nothing here assumes any particular heading exists.
    """
    grid = _read_grid(path, sheet_name)
    if len(grid) < header_row:
        raise RuntimeError(
            f"That sheet has no row {header_row}, so there are no column names to read."
        )

    headers: list[str] = []
    seen: set[str] = set()
    duplicates: list[str] = []
    index_for: dict[str, int] = {}
    for i, cell in enumerate(grid[header_row - 1]):
        title = cell_to_text(cell)
        if not title:
            continue
        key = norm(title)
        if key in seen:
            duplicates.append(title)
            continue
        seen.add(key)
        headers.append(title)
        index_for[key] = i

    if not headers:
        raise RuntimeError(
            f"Row {header_row} of that sheet is empty, so the app cannot tell what the "
            "columns are. Check which row holds your column names."
        )

    rows: list[BatchRow] = []
    blank_run = 0
    for number, raw in enumerate(grid[header_row:], start=header_row + 1):
        values = {
            key: cell_to_text(raw[i]) if i < len(raw) else ""
            for key, i in index_for.items()
        }
        if not any(values.values()):
            # A formatted-but-empty workbook can claim a million rows. Read past a few
            # gaps in the middle of a list, but stop rather than churn through the lot.
            blank_run += 1
            if blank_run >= MAX_BLANK_RUN:
                break
            continue
        blank_run = 0
        rows.append(BatchRow(row=number, values=values))

    return Table(headers=headers, rows=rows, duplicate_headers=duplicates)


def _read_grid(path: Path, sheet_name: str | None) -> list[tuple]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            return [tuple(r) for r in csv.reader(fh)]

    from openpyxl import load_workbook

    # data_only: a reference number produced by a formula has to arrive as its value, not
    # as "=CONCAT(...)". The cost is that a workbook Excel has never opened and saved has
    # no cached values, which read_table surfaces as empty cells rather than pretending.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        want = pd_lib.find_sheet(names, sheet_name) if sheet_name else None
        ws = wb[want or names[0]]
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def sheet_names(path: Path) -> tuple[list[str], str | None]:
    """Visible sheet names, plus whichever one Excel had open when it was saved."""
    if path.suffix.lower() == ".csv":
        return [], None
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        try:
            active = wb.active.title if wb.active is not None else None
        except Exception:  # noqa: BLE001 - a stale active index shouldn't stop the upload
            active = None
        return (visible or list(wb.sheetnames)), active
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Matching placeholders to columns
# ---------------------------------------------------------------------------

def guess_column(keys: list[str], name: str) -> str:
    """The column a placeholder most likely means, or "" when it is not obvious.

    Three passes, each accepted only when it is unambiguous: the name as written, the name
    with spacing and punctuation thrown away ("Ref No." finding "Ref no"), then one
    containing the other ("{{Reference}}" finding "Reference Number", but only if exactly
    one column could be meant). Guessing wrong is worse than not guessing at all, because a
    wrong guess is silent while an unmatched placeholder puts a dropdown on screen.
    """
    key = norm(name)
    if key in keys:
        return key

    squashed = _squash(name)
    if not squashed:
        return ""

    hits = [k for k in keys if _squash(k) == squashed]
    if len(hits) == 1:
        return hits[0]

    hits = [k for k in keys if squashed in _squash(k) or _squash(k) in squashed]
    return hits[0] if len(hits) == 1 else ""


def guess_email_column(keys: list[str]) -> str:
    """The column holding the recipient address, or "" if nothing looks like one."""
    for exact in ("email address", "e-mail address", "email", "e-mail", "recipient email"):
        if exact in keys:
            return exact
    hits = [k for k in keys if "email" in k or "e-mail" in k]
    if len(hits) == 1:
        return hits[0]
    # Prefer a real address column over an "Email Sent?" style bookkeeping one.
    for k in hits:
        if not re.search(r"\b(sent|date|status|opt|bounce|valid)\b", k):
            return k
    return hits[0] if hits else ""


def auto_mapping(placeholders: list[str], keys: list[str]) -> dict[str, str]:
    """Normalised placeholder -> normalised column, for everything that matched."""
    out: dict[str, str] = {}
    for name in placeholders:
        guess = guess_column(keys, name)
        if guess:
            out[norm(name)] = guess
    return out


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def render(text: str, values: dict[str, str], mapping: dict[str, str]) -> tuple[str, list[str]]:
    """Substitute this row's values into a template.

    Returns the rendered text and the placeholders that had nothing to put there. An
    unresolved placeholder is left standing as {{Reference Number}} on purpose: the row is
    blocked either way, and leaving the marker means the preview shows exactly where the
    hole is, instead of a plausible-looking sentence with a word quietly missing.

    A placeholder written with a trailing question mark — {{Job Title?}} — is optional: an
    empty cell renders as nothing and does not block the row. Blank is the strict default
    because that is the dangerous case, but some things genuinely do not apply to everyone,
    and without this the only way to send them is to split the batch.
    """
    missing: list[str] = []

    def replace(m: re.Match) -> str:
        raw = m.group(1).strip()
        column = mapping.get(norm(raw), "")
        value = values.get(column, "") if column else ""
        if not value:
            if raw.endswith("?"):
                return ""
            if raw not in missing:
                missing.append(raw)
            return m.group(0)
        return value

    return PLACEHOLDER_RE.sub(replace, text or ""), missing


def tidy_body(text: str) -> str:
    """Close up the gaps a merge leaves behind.

    Templates often have a line that is nothing but a placeholder — a reference, a job
    title on its own line. Where that line is legitimately blank for a row it should
    disappear, not leave a stranded empty line in the middle of the email.
    """
    lines = [re.sub(r"[ \t]+", " ", l).rstrip() for l in (text or "").splitlines()]
    kept = [l for i, l in enumerate(lines) if l or (i and lines[i - 1].strip())]
    return re.sub(r"\n{3,}", "\n\n", "\n".join(kept)).strip()


@dataclass
class RowDraft:
    """One row, rendered. `problems` being empty is the only thing that lets it be sent."""
    row: int
    to: str
    subject: str
    body: str
    missing: list[str] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    label: str = ""

    @property
    def ok(self) -> bool:
        return not self.problems


def build_drafts(
    table: Table,
    subject_template: str,
    body_template: str,
    mapping: dict[str, str],
    email_column: str,
    label_column: str = "",
) -> list[RowDraft]:
    """Render every row, and say plainly which ones are not fit to send.

    Nothing is filtered out here. A blocked row still comes back, carrying the reason, so
    the app can show it by row number and the user can go and fix the spreadsheet. Silently
    dropping thirteen of a hundred rows and reporting eighty-seven successes is how a batch
    goes out short without anybody noticing.
    """
    drafts: list[RowDraft] = []
    for row in table.rows:
        to = row.values.get(email_column, "").strip()
        subject, missing_subject = render(subject_template, row.values, mapping)
        body, missing_body = render(body_template, row.values, mapping)
        body = tidy_body(body)
        subject = re.sub(r"\s+", " ", subject).strip()

        missing = list(dict.fromkeys(missing_subject + missing_body))
        problems: list[str] = []
        if not to:
            problems.append("No email address")
        elif not EMAIL_RE.match(to):
            problems.append(f"“{to}” is not a valid email address")
        for name in missing:
            problems.append(f"Nothing in the “{name}” column")
        if not subject:
            problems.append("The subject line came out empty")
        if not body:
            problems.append("The email came out empty")

        label = row.values.get(label_column, "").strip() if label_column else ""
        drafts.append(
            RowDraft(
                row=row.row, to=to, subject=subject, body=body,
                missing=missing, problems=problems,
                label=label or to or f"Row {row.row}",
            )
        )

    _flag_duplicates(drafts)
    return drafts


def _flag_duplicates(drafts: list[RowDraft]) -> None:
    """Warn where the same address appears twice — the same person, two rows, two emails.

    A warning rather than a block: a shared inbox legitimately appears several times in
    some lists, and this is a judgement the user has to make, not the app.
    """
    seen: dict[str, list[int]] = {}
    for d in drafts:
        if d.to:
            seen.setdefault(d.to.lower(), []).append(d.row)
    for d in drafts:
        rows = seen.get(d.to.lower(), [])
        if len(rows) > 1:
            others = ", ".join(str(r) for r in rows if r != d.row)
            d.warnings.append(f"This address is also on row {others}")


# ---------------------------------------------------------------------------
# Word templates
# ---------------------------------------------------------------------------

_DOCX_PARA_RE = re.compile(r"<w:p\b[^>]*/>|<w:p\b[^>]*>(.*?)</w:p>", re.S)
_DOCX_TOKEN_RE = re.compile(
    r"<w:t(?:\s[^>]*)?>(.*?)</w:t>|<w:tab\b[^>]*>|<w:br\b[^>]*>|<w:cr\b[^>]*>", re.S
)

# Word stores a hyperlink in two halves: the paragraph carries <w:hyperlink r:id="rId7">
# around the words you see, and the target itself lives in a separate relationships file
# keyed by that id. Reading only document.xml is why the address went missing.
_DOCX_HYPERLINK_RE = re.compile(r"<w:hyperlink\b([^>]*)>(.*?)</w:hyperlink\s*>", re.S)
_DOCX_REL_RE = re.compile(
    r"<Relationship\b[^>]*Id=\"([^\"]+)\"[^>]*Target=\"([^\"]*)\"[^>]*>", re.I
)
_DOCX_REL_REVERSED_RE = re.compile(
    r"<Relationship\b[^>]*Target=\"([^\"]*)\"[^>]*Id=\"([^\"]+)\"[^>]*>", re.I
)
_DOCX_RID_RE = re.compile(r"r:id=\"([^\"]+)\"", re.I)

# The older field-code form of the same thing, which Word still writes in places:
#   <w:instrText> HYPERLINK "https://..." </w:instrText> ... <w:fldChar w:fldCharType="end"/>
_DOCX_FIELD_LINK_RE = re.compile(
    r"<w:instrText[^>]*>\s*HYPERLINK\s+\"([^\"]+)\"[^<]*</w:instrText>"
    r"(.*?)<w:fldChar[^>]*w:fldCharType=\"end\"[^>]*>",
    re.S | re.I,
)


def _xml_escape(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _docx_runs_text(xml: str) -> str:
    pieces: list[str] = []
    for t in _DOCX_TOKEN_RE.finditer(xml):
        if t.group(1) is not None:
            pieces.append(html.unescape(t.group(1)))
        elif t.group(0).startswith("<w:tab"):
            pieces.append(" ")
        else:
            pieces.append("\n")
    return re.sub(r"[ \t]+", " ", "".join(pieces)).strip()


def _docx_relationships(z: zipfile.ZipFile) -> dict[str, str]:
    try:
        rels = z.read("word/_rels/document.xml.rels").decode("utf-8", "replace")
    except KeyError:
        return {}
    out = {rid: target for rid, target in _DOCX_REL_RE.findall(rels)}
    # Attribute order is not guaranteed, and a Relationship written Target-before-Id would
    # otherwise be missed, taking its link with it.
    for target, rid in _DOCX_REL_REVERSED_RE.findall(rels):
        out.setdefault(rid, target)
    return out


def _inline_docx_links(xml: str, rels: dict[str, str]) -> str:
    """Rewrite each hyperlink as a run of ordinary text carrying its target alongside.

    Folding links into the run text before the paragraph is walked means the rest of the
    conversion needs to know nothing about them, and a link is picked up wherever Word put
    it — mid-sentence, in a table cell, inside a bullet.
    """
    def as_run(label: str, target: str) -> str:
        label, target = label.strip(), (target or "").strip()
        if not target:
            return f"<w:t>{_xml_escape(label)}</w:t>"
        written = target if (not label or label == target) else f"[{label}]({target})"
        return f"<w:t>{_xml_escape(written)}</w:t>"

    def replace_anchor(m: re.Match) -> str:
        rid = _DOCX_RID_RE.search(m.group(1) or "")
        target = rels.get(rid.group(1), "") if rid else ""
        label = _docx_runs_text(m.group(2))
        # A w:hyperlink with no r:id is an internal jump to a bookmark in the same document.
        # There is nothing to link to in an email, so only the words survive.
        return as_run(label, target)

    xml = _DOCX_HYPERLINK_RE.sub(replace_anchor, xml)
    return _DOCX_FIELD_LINK_RE.sub(
        lambda m: as_run(_docx_runs_text(m.group(2)), m.group(1)), xml
    )


def docx_to_text(data: bytes) -> str:
    """The paragraph text of a .docx, with a blank line between paragraphs.

    Hyperlinks survive, written as [visible words](https://the-target) — see the Links
    section above. Everything else about the formatting is dropped: bold, colour, font and
    images all go, and a table comes out as its cell text one line at a time. That is a real
    limitation and the app says so on screen rather than letting someone discover it in a
    sent email.

    Done by reading the document XML directly instead of adding a Word library to
    requirements.txt: this app is deployed by someone else through an IT request, and one
    fewer dependency to justify is worth more here than the rest of the formatting would be.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            xml = z.read("word/document.xml").decode("utf-8", "replace")
            xml = _inline_docx_links(xml, _docx_relationships(z))
    except KeyError as exc:  # noqa: PERF203 - one clear message beats a stack trace
        raise RuntimeError(
            "That .docx has no document body the app can read. Try opening it in Word and "
            "using Save As to write a fresh copy."
        ) from exc
    except zipfile.BadZipFile as exc:
        raise RuntimeError(
            "That does not look like a Word .docx file. If it is an older .doc, open it in "
            "Word and save it as .docx first — or just paste the email in instead."
        ) from exc

    paragraphs: list[tuple[str, bool]] = []
    for m in _DOCX_PARA_RE.finditer(xml):
        inner = m.group(1) or ""
        # A numbered or bulleted paragraph loses its marker along with the formatting, so
        # put a plain dash back — a bare run of lines reads as one run-on paragraph.
        bullet = "<w:numPr" in inner
        text = _docx_runs_text(inner)
        paragraphs.append((("- " + text) if (text and bullet) else text, bullet and bool(text)))

    # Blocks are joined with a blank line, so Word's own empty paragraphs are simply
    # dropped: paragraph spacing comes from the join. A run of bullets is one block, or a
    # three-item list would arrive as three separate paragraphs.
    blocks: list[str] = []
    prev_bullet = False
    for text, bullet in paragraphs:
        if not text:
            prev_bullet = False
            continue
        if bullet and prev_bullet and blocks:
            blocks[-1] += "\n" + text
        else:
            blocks.append(text)
        prev_bullet = bullet

    return re.sub(r"\n{3,}", "\n\n", "\n\n".join(blocks)).strip()


def read_template_upload(name: str, data: bytes) -> str:
    """A .docx, .md or .txt upload as the plain text of an email."""
    suffix = Path(name).suffix.lower()
    if suffix == ".docx":
        return docx_to_text(data)
    if suffix in (".md", ".txt", ""):
        return data.decode("utf-8", "replace").replace("\r\n", "\n").strip()
    raise RuntimeError(f"The app cannot read a {suffix} file. Use .docx, .txt or .md.")


# ---------------------------------------------------------------------------
# The log
# ---------------------------------------------------------------------------

def log_csv(rows: list[dict]) -> bytes:
    """A record of what this batch did, one line per row of the list.

    Worth having for its own sake, and it is also what makes a second run safe: it is the
    only way to tell, next week, which of the hundred already went out.
    """
    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow(["Sheet row", "To", "Subject", "Outcome", "Detail"])
    for r in rows:
        w.writerow([
            r.get("row", ""), r.get("to", ""), r.get("subject", ""),
            r.get("outcome", ""), r.get("detail", ""),
        ])
    return buf.getvalue().encode("utf-8-sig")
